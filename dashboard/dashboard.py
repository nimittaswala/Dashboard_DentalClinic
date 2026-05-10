# -*- coding: utf-8 -*-
"""
Dental Automation Control Panel
Run: python dashboard.py
Open: http://127.0.0.1:5000
"""

import sys, os, json, logging, threading, subprocess, smtplib
from datetime import datetime, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from flask import Flask, jsonify, request, session, redirect, render_template_string
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger

# =============================================================================
# CONFIG
# =============================================================================

sys.path.insert(0, r"C:\dental_automation\config")

try:
    from office_config import OFFICE, GMAIL_CONFIG, MANAGER_EMAIL, FOLDERS, DB_CONFIG
except ImportError as e:
    print(f"ERROR loading office_config.py: {e}")
    sys.exit(1)

# Optional dashboard password (backward compat — if not set, no auth)
try:
    from office_config import DASHBOARD_PASSWORD
    DASHBOARD_PASSWORD = (DASHBOARD_PASSWORD or "").strip() or None
except ImportError:
    DASHBOARD_PASSWORD = None

PRACTICE_NAME = OFFICE["name"]
BASE_DIR      = FOLDERS["base"]
LOG_FILE      = os.path.join(BASE_DIR, "dashboard", "dashboard.log")
os.makedirs(os.path.dirname(LOG_FILE), exist_ok=True)

logging.basicConfig(
    filename=LOG_FILE, level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S"
)

def log(msg):
    print(msg)
    logging.info(msg)

# =============================================================================
# SCRIPT DEFINITIONS
# =============================================================================

SCRIPTS = {
    "unsigned_notes": {
        "name": "Unsigned Notes",
        "desc": "Daily provider note reminders with auto-note templates",
        "path": os.path.join(BASE_DIR, "unsigned_notes", "unsigned_notes.py"),
        "icon": "[Notes]",
        "color": "#3498db",
        "cron": "Mon-Sat at 10:00 AM",
        "schedule": {"day_of_week": "mon-sat", "hour": 10, "minute": 0},
    },
    "huddle_summary": {
        "name": "Huddle Summary",
        "desc": "Morning huddle - schedule, unconfirmed, insurance verify",
        "path": os.path.join(BASE_DIR, "huddle_summary", "huddle_summary.py"),
        "icon": "[Huddle]",
        "color": "#e67e22",
        "cron": "Mon-Sat at 8:00 AM",
        "schedule": {"day_of_week": "mon-sat", "hour": 8, "minute": 0},
    },
    "recall_list": {
        "name": "Recall List",
        "desc": "Weekly overdue recall patients with probability scores",
        "path": os.path.join(BASE_DIR, "recall_list", "recall_list.py"),
        "icon": "[Recall]",
        "color": "#9b59b6",
        "cron": "Monday at 8:00 AM",
        "schedule": {"day_of_week": "mon", "hour": 8, "minute": 0},
    },
    "production": {
        "name": "Production Dashboard",
        "desc": "End of day - patients, procedures, collections",
        "path": os.path.join(BASE_DIR, "production", "production_dashboard.py"),
        "icon": "[Production]",
        "color": "#2ecc71",
        "cron": "Mon-Sat at 7:00 PM",
        "schedule": {"day_of_week": "mon-sat", "hour": 19, "minute": 0},
    },
    "ar_push": {
        "name": "AR Push",
        "desc": "Push outstanding insurance AR to central dashboard",
        "path": os.path.join(BASE_DIR, "ar", "ar_push.py"),
        "icon": "[AR]",
        "color": "#e74c3c",
        "cron": "Daily at 7:00 PM",
        "schedule": {"hour": 19, "minute": 0},
    },
}

# =============================================================================
# STATE
# =============================================================================

script_state = {
    sid: {
        "paused": False, "last_run": "",
        "last_status": "", "last_summary": "",
        "running": False, "next_run": "",
    }
    for sid in SCRIPTS
}

activity_log = []

# =============================================================================
# SCHEDULER
# =============================================================================

scheduler = BackgroundScheduler()

def run_script(script_id):
    if script_state[script_id]["paused"]:
        return
    script_state[script_id]["running"] = True
    script = SCRIPTS[script_id]
    start  = datetime.now()
    log(f"Running: {script['name']}")
    try:
        result  = subprocess.run(
            [sys.executable, script["path"]],
            capture_output=True, text=True, timeout=300
        )
        success = result.returncode == 0
        output  = result.stdout + result.stderr
        status  = "Success" if success else "Failed"
        summary = parse_summary(output)
        script_state[script_id].update({
            "last_run": start.strftime("%I:%M %p"),
            "last_status": status,
            "last_summary": summary,
            "running": False,
        })
        activity_log.insert(0, {
            "time": start.strftime("%I:%M %p"),
            "script": script["name"],
            "icon": script["icon"],
            "status": status,
            "summary": summary,
            "color": "#2ecc71" if success else "#e74c3c",
        })
        if len(activity_log) > 50:
            activity_log.pop()
        if not success:
            send_failure_alert(script["name"], output)
    except Exception as e:
        script_state[script_id].update({
            "last_run": start.strftime("%I:%M %p"),
            "last_status": "Error",
            "last_summary": str(e),
            "running": False,
        })
        send_failure_alert(script["name"], str(e))
    update_next_runs()

def parse_summary(output):
    lines = output.strip().split("\n")
    keep  = []
    for line in lines:
        line = line.strip()
        if any(kw in line for kw in ["Found","sent to","Error","patients","unsigned","Collected"]):
            if "[INFO]" in line:
                line = line.split("[INFO]")[-1].strip()
            keep.append(line)
    return " | ".join(keep[:4]) if keep else "Completed"

def update_next_runs():
    for sid in SCRIPTS:
        if script_state[sid]["paused"]:
            script_state[sid]["next_run"] = "Paused"
            continue
        job = scheduler.get_job(sid)
        if job:
            try:
                nr = getattr(job, "next_run_time", None) or getattr(job, "next_fire_time", None)
                script_state[sid]["next_run"] = nr.strftime("%I:%M %p %A") if nr else "Scheduled"
            except Exception:
                script_state[sid]["next_run"] = "Scheduled"
        else:
            script_state[sid]["next_run"] = "Not scheduled"

def send_failure_alert(script_name, error):
    if not MANAGER_EMAIL:
        return
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = f"Alert: {script_name} Failed | {PRACTICE_NAME}"
        msg["From"]    = GMAIL_CONFIG["sender_email"]
        msg["To"]      = MANAGER_EMAIL
        body = f"<p><b>{script_name}</b> failed at {datetime.now().strftime('%I:%M %p')}.</p><pre>{error[-1000:]}</pre>"
        msg.attach(MIMEText(body, "html"))
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
            s.login(GMAIL_CONFIG["sender_email"], GMAIL_CONFIG["sender_password"])
            s.sendmail(GMAIL_CONFIG["sender_email"], MANAGER_EMAIL, msg.as_string())
    except Exception:
        pass

def add_jobs():
    for sid, script in SCRIPTS.items():
        sched = script["schedule"]
        if "day_of_week" in sched:
            trigger = CronTrigger(day_of_week=sched["day_of_week"],
                                   hour=sched["hour"], minute=sched["minute"])
        else:
            trigger = CronTrigger(hour=sched["hour"], minute=sched["minute"])
        scheduler.add_job(run_script, trigger, args=[sid], id=sid, replace_existing=True, misfire_grace_time=300)
    update_next_runs()

# =============================================================================
# CLINICAL SNAPSHOT — Database queries and logic
# =============================================================================

import math
from datetime import date, time as dtime

# Medicaid carriers — no copay flag
MEDICAID_CARRIERS = [
    "keystone first", "keystone mercy", "keystone first dentaquest",
    "keystone first vip", "health partner", "health partners",
    "health partner medicare", "aetna better health",
    "united healthcare community plan", "upmc", "avesis", "dentaquest",
]

# Providers excluded from chair time / health score
NO_CHAIR_TIME_PROVIDERS = [
    "Ankita Lulla", "Maha Tayefeh", "Dhaval Shah", "Nesreen Ibrahim"
]

# High value procedure codes
HIGH_VALUE_CODES = {"D2740","D2700","D3330","D3320","D3310","D6010","D6065",
                    "D6066","D4341","D4342","D7210","D5110","D5120"}
MEDIUM_VALUE_CODES = {"D2391","D2392","D2393","D2331","D2332","D2335",
                      "D7140","D7953","D4910"}
RECALL_CODES = {"D0120","D1110","D1120","D0150","D0140","D0220","D0230",
                "D0274","D0330","D1208","D1351","D1310","D1330","D0601"}
PRIOR_AUTH_CODES = {"D4341","D4342","D2740","D5110","D5120"}
COPAY_TRIGGER_CODES = {"D2391","D2392","D2393","D2331","D2332","D2335",
                       "D2740","D2700","D4341","D4342","D7140","D7210",
                       "D7953","D3330","D3320","D3310","D6010","D5110",
                       "D5120","D5201","D5202","D5203"}

def get_snapshot_date():
    """Return the display date — Monday if today is Sunday."""
    today = date.today()
    if today.weekday() == 6:  # Sunday
        from datetime import timedelta
        return today + timedelta(days=1), True
    return today, False

def minutes_since_midnight(t):
    """Convert time object to minutes since midnight."""
    return t.hour * 60 + t.minute

def get_snapshot_data(cursor, target_date=None):
    """Fetch all data needed for clinical snapshot."""
    if target_date is not None:
        snap_date = target_date
        is_preview = True
    else:
        snap_date, is_preview = get_snapshot_date()

    # -------------------------------------------------------------------------
    # 1. Provider schedules for today
    # -------------------------------------------------------------------------
    cursor.execute("""
        SELECT s.ProvNum, s.StartTime, s.StopTime, s.SchedType,
               s.BlockoutType, s.Note,
               CONCAT(p.FName, ' ', p.LName) AS ProviderName,
               p.Abbr
        FROM schedule s
        LEFT JOIN provider p ON p.ProvNum = s.ProvNum
        WHERE s.SchedDate = %s
        AND s.SchedType IN ('Provider', 'Blockout')
        ORDER BY s.SchedType, s.StartTime
    """, (snap_date,))
    schedules = cursor.fetchall()

    prov_schedules = {}   # ProvNum -> {start, stop, name, abbr}
    blockouts      = []   # list of {start_min, stop_min}

    for s in schedules:
        if s["SchedType"] == "Provider" and s["ProvNum"]:
            pn = s["ProvNum"]
            if pn not in prov_schedules:
                prov_schedules[pn] = {
                    "name":      s["ProviderName"] or "Unknown",
                    "abbr":      s["Abbr"] or "",
                    "start_min": minutes_since_midnight(s["StartTime"]),
                    "stop_min":  minutes_since_midnight(s["StopTime"]),
                    "start_str": s["StartTime"].strftime("%I:%M %p").lstrip("0"),
                    "stop_str":  s["StopTime"].strftime("%I:%M %p").lstrip("0"),
                }
        elif s["SchedType"] == "Blockout":
            blockouts.append({
                "start_min": minutes_since_midnight(s["StartTime"]),
                "stop_min":  minutes_since_midnight(s["StopTime"]),
                "note":      s["Note"] or "Blockout",
            })

    # -------------------------------------------------------------------------
    # 2. Today's appointments with full patient/insurance info
    # -------------------------------------------------------------------------
    cursor.execute("""
        SELECT
            a.AptNum, a.AptDateTime, a.AptStatus, a.Confirmed,
            COALESCE(def.ItemName, 'Unknown') AS ConfirmedName,
            a.ProvNum, a.Op, a.Pattern,
            a.ProcDescript,
            CONCAT(pat.LName, ', ', pat.FName) AS PatientName,
            COALESCE(NULLIF(pat.WirelessPhone,''),
                     NULLIF(pat.HmPhone,''),
                     NULLIF(pat.WkPhone,'')) AS Phone,
            pat.Birthdate, pat.Premed, pat.MedUrgNote,
            pat.BalTotal, pat.DateFirstVisit, pat.PatNum,
            CONCAT(prov.FName, ' ', prov.LName) AS ProviderName,
            prov.Abbr AS ProvAbbr,
            op.OpName,
            pp.InsSubNum,
            COALESCE(ip.PlanType, '') AS PlanType,
            c.CarrierName,
            iv.DateLastVerified,
            af.FieldValue AS CoPay,
            pat.DateFirstVisit
        FROM appointment a
        INNER JOIN patient pat ON pat.PatNum = a.PatNum
        INNER JOIN provider prov ON prov.ProvNum = a.ProvNum
        LEFT JOIN operatory op ON op.OperatoryNum = a.Op
        LEFT JOIN patplan pp ON pp.PatNum = a.PatNum
            AND pp.Ordinal = 1 AND pp.IsPending = 0
        LEFT JOIN inssub ins ON ins.InsSubNum = pp.InsSubNum
        LEFT JOIN insplan ip ON ip.PlanNum = ins.PlanNum
        LEFT JOIN carrier c ON c.CarrierNum = ip.CarrierNum
        LEFT JOIN (
            SELECT pp_inner.InsSubNum,
                MAX(iv_all.DateLastVerified) AS DateLastVerified
            FROM patplan pp_inner
            INNER JOIN (
                SELECT FKey, VerifyType, DateLastVerified FROM insverify
                WHERE DateLastVerified > '0001-01-01'
                UNION ALL
                SELECT FKey, VerifyType, DateLastVerified FROM insverifyhist
                WHERE DateLastVerified > '0001-01-01'
            ) iv_all ON (
                (iv_all.VerifyType = 1 AND iv_all.FKey = pp_inner.InsSubNum)
                OR
                (iv_all.VerifyType = 2 AND iv_all.FKey = pp_inner.PatPlanNum)
            )
            GROUP BY pp_inner.InsSubNum
        ) iv ON iv.InsSubNum = pp.InsSubNum
        LEFT JOIN apptfield af ON af.AptNum = a.AptNum
            AND af.FieldName = 'CoPay'
        LEFT JOIN definition def ON def.DefNum = a.Confirmed
        WHERE a.AptStatus IN (1, 2, 3, 5)
        AND DATE(a.AptDateTime) = %s
        ORDER BY a.AptDateTime ASC
    """, (snap_date,))
    appointments = cursor.fetchall()

    # -------------------------------------------------------------------------
    # 3. Procedures for today's appointments
    # -------------------------------------------------------------------------
    if appointments:
        apt_nums = [a["AptNum"] for a in appointments]
        fmt      = ",".join(["%s"] * len(apt_nums))
        cursor.execute(f"""
            SELECT pl.AptNum, pc.ProcCode, pc.Descript
            FROM procedurelog pl
            INNER JOIN procedurecode pc ON pc.CodeNum = pl.CodeNum
            WHERE pl.AptNum IN ({fmt})
            AND pl.ProcStatus IN (1, 2)
        """, apt_nums)
        proc_rows = cursor.fetchall()
    else:
        proc_rows = []

    apt_procs = {}
    for row in proc_rows:
        apt_procs.setdefault(row["AptNum"], []).append(row["ProcCode"])

    # -------------------------------------------------------------------------
    # 4. No-show history for today's patients
    # -------------------------------------------------------------------------
    noshow_history = {}
    if appointments:
        pat_nums = list(set(a["PatNum"] for a in appointments))
        fmt      = ",".join(["%s"] * len(pat_nums))
        cursor.execute(f"""
            SELECT PatNum,
                COUNT(*) AS TotalApts,
                SUM(CASE WHEN AptStatus = 'Broken' THEN 1 ELSE 0 END) AS Broken,
                SUM(CASE WHEN AptStatus = 'UnschedList' THEN 1 ELSE 0 END) AS Unscheduled
            FROM appointment
            WHERE PatNum IN ({fmt})
            AND AptDateTime < NOW()
            GROUP BY PatNum
        """, pat_nums)
        for row in cursor.fetchall():
            total   = row["TotalApts"] or 1
            bad     = (row["Broken"] or 0) + (row["Unscheduled"] or 0)
            rate    = bad / total
            risk    = "high" if bad >= 3 else "medium" if bad >= 1 else "low"
            noshow_history[row["PatNum"]] = {
                "bad": bad, "total": total,
                "rate": rate, "risk": risk
            }

    # -------------------------------------------------------------------------
    # 5. Last visit date for today's patients
    # -------------------------------------------------------------------------
    last_visit = {}
    if appointments:
        pat_nums = list(set(a["PatNum"] for a in appointments))
        fmt      = ",".join(["%s"] * len(pat_nums))
        cursor.execute(f"""
            SELECT PatNum, MAX(AptDateTime) AS LastVisit
            FROM appointment
            WHERE PatNum IN ({fmt})
            AND AptStatus IN ('Complete', 2)
            AND DATE(AptDateTime) < %s
            GROUP BY PatNum
        """, pat_nums + [snap_date])
        for row in cursor.fetchall():
            if row["LastVisit"]:
                delta = (snap_date - row["LastVisit"].date()).days
                last_visit[row["PatNum"]] = delta

    # -------------------------------------------------------------------------
    # 5b. Previous provider for today's patients
    # -------------------------------------------------------------------------
    prev_provider = {}
    if appointments:
        pat_nums = list(set(a["PatNum"] for a in appointments))
        fmt      = ",".join(["%s"] * len(pat_nums))
        prev_sql = (
            "SELECT pl.PatNum, "
            "CONCAT(prov.FName,' ',prov.LName) AS PrevProvider, "
            "pl.ProvNum AS PrevProvNum "
            "FROM procedurelog pl "
            "INNER JOIN provider prov ON prov.ProvNum = pl.ProvNum "
            "WHERE pl.PatNum IN (" + fmt + ") "
            "AND pl.ProcStatus = 2 "
            "AND pl.ProcDate < %s "
            "AND pl.ProcDate = ("
            "  SELECT MAX(pl2.ProcDate) FROM procedurelog pl2 "
            "  WHERE pl2.PatNum = pl.PatNum AND pl2.ProcStatus = 2 "
            "  AND pl2.ProcDate < %s) "
            "GROUP BY pl.PatNum, pl.ProvNum, PrevProvider"
        )
        cursor.execute(prev_sql, pat_nums + [snap_date, snap_date])
        for row in cursor.fetchall():
            pn = row["PatNum"]
            if pn not in prev_provider:
                prev_provider[pn] = {
                    "name":     row["PrevProvider"],
                    "prov_num": row["PrevProvNum"]
                }

    # -------------------------------------------------------------------------
    # 5c. Routing slips scanned today
    # Dynamically find the DocCategory DefNum by searching definition table
    # for any folder whose name contains 'routing' — works across servers
    # even if the folder is named differently or has a different DefNum.
    # -------------------------------------------------------------------------
    routing_slip_pats = set()
    if appointments:
        # Find routing slip category DefNum dynamically
        cursor.execute("""
            SELECT DefNum FROM definition
            WHERE LOWER(ItemName) LIKE '%routing%'
            AND Category = 18
            LIMIT 1
        """)
        cat_row = cursor.fetchone()
        routing_cat = cat_row["DefNum"] if cat_row else None

        if routing_cat:
            pat_nums = list(set(a["PatNum"] for a in appointments))
            fmt      = ",".join(["%s"] * len(pat_nums))
            cursor.execute(f"""
                SELECT DISTINCT PatNum
                FROM document
                WHERE DocCategory = %s
                AND DATE(DateCreated) = %s
                AND PatNum IN ({fmt})
            """, [routing_cat, snap_date] + pat_nums)
            for row in cursor.fetchall():
                routing_slip_pats.add(row["PatNum"])
    # -------------------------------------------------------------------------
    today_str  = snap_date.strftime("%A, %B %d, %Y")
    snap_date_str = snap_date.isoformat()
    enriched   = []
    alerts     = {"unconfirmed": [], "ins_verify": [],
                  "copay": [], "prior_auth": []}

    NEEDS_CALL = ("Unconfrm", "LeftMsg", "MsgFam")
    today_md   = (snap_date.month, snap_date.day)

    for apt in appointments:
        pat_num   = apt["PatNum"]
        codes     = apt_procs.get(apt["AptNum"], [])
        codes_set = set(codes)
        nh        = noshow_history.get(pat_num, {"bad":0,"total":0,"rate":0,"risk":"low"})
        lv        = last_visit.get(pat_num)
        lv_str    = f"{lv}d ago" if lv is not None else "First visit"
        carrier   = apt["CarrierName"] or ""
        plan_type = apt["PlanType"] or ""
        is_medicaid = (plan_type == "f" or
                       any(m in carrier.lower() for m in MEDICAID_CARRIERS))
        has_ins   = bool(apt["InsSubNum"])
        is_recall_only = codes_set and codes_set.issubset(RECALL_CODES)
        time_str  = apt["AptDateTime"].strftime("%I:%M %p").lstrip("0")
        dob       = apt["Birthdate"]

        # Resolve confirmed status name from definition table
        conf_name = apt.get("ConfirmedName") or str(apt.get("Confirmed", ""))

        row = {
            "aptnum":      apt["AptNum"],
            "time":        time_str,
            "time_min":    minutes_since_midnight(apt["AptDateTime"].time()),
            "patient":     apt["PatientName"],
            "pat_num":     pat_num,
            "provider":    apt["ProviderName"],
            "prov_num":    apt["ProvNum"],
            "prov_abbr":   apt["ProvAbbr"],
            "room":        apt["OpName"] or "",
            "confirmed":   conf_name,
            "duration_min": len(apt.get("Pattern") or "") * 5 or 30,
            "codes":       codes,
            "copay":       apt["CoPay"] or "",
            "carrier":     carrier,
            "plan_type":   plan_type,
            "is_medicaid": is_medicaid,
            "has_ins":     has_ins,
            "last_visit":  lv_str,
            "noshow_risk": nh["risk"],
            "noshow_bad":  nh["bad"],

            "phone":       apt["Phone"] or "",
            "is_new":      (apt["DateFirstVisit"] and
                           apt["DateFirstVisit"] == snap_date),
            "prev_prov":   prev_provider.get(pat_num, {}).get("name", ""),
            "prov_changed": (
                pat_num in prev_provider and
                prev_provider[pat_num]["prov_num"] != apt["ProvNum"]
            ),
            "apt_status":  int(apt["AptStatus"] or 1),
            "routing_slip": pat_num in routing_slip_pats,
        }
        enriched.append(row)

        # Skip alert generation for broken/cancelled/unscheduled
        if int(apt.get("AptStatus") or 1) in (3, 5):
            continue

        # --- Alerts ---
        CONFIRMED_OK = (
            "Appointment Confirmed", "Confirmed",
            "Arrived", "In Treatment Room", "In Room",
            "Ready", "Claim Sent", "No Bill to Insurance",
            "No Bill", "Complete"
        )
        if conf_name not in CONFIRMED_OK:
            alerts["unconfirmed"].append({
                "time": time_str, "patient": apt["PatientName"],
                "provider": apt["ProviderName"], "phone": apt["Phone"] or "No phone",
                "status": conf_name or "Unconfirmed"
            })

        if has_ins:
            dlv = apt["DateLastVerified"]
            # Treat null, 0001-01-01, or empty as never verified
            never = (not dlv or str(dlv).startswith("0001"))
            if never:
                color = "red"
                label = "Never verified"
            else:
                try:
                    from datetime import date as _date, datetime as _datetime
                    if isinstance(dlv, _datetime):
                        dlv_date = dlv.date()
                    elif isinstance(dlv, _date):
                        dlv_date = dlv
                    else:
                        dlv_date = _datetime.strptime(str(dlv), "%Y-%m-%d").date()
                    if str(dlv_date).startswith("0001"):
                        color = "red"
                        label = "Never verified"
                        dlv_date = None
                    else:
                        days_since = (snap_date - dlv_date).days
                    if days_since <= 7:
                        color = "green"
                        label = f"{days_since}d ago"
                    elif days_since <= 15:
                        color = "yellow"
                        label = f"{days_since}d ago"
                    else:
                        color = "red"
                        label = f"{days_since}d ago"
                except Exception:
                    color = "red"
                    label = "Check verify date"
            if color != "green":
                alerts["ins_verify"].append({
                    "time": time_str, "patient": apt["PatientName"],
                    "provider": apt["ProviderName"], "carrier": carrier,
                    "label": label, "color": color
                })

        copay_val = apt["CoPay"] or ""
        if copay_val:
            alerts["copay"].append({
                "time": time_str, "patient": apt["PatientName"],
                "provider": apt["ProviderName"], "amount": copay_val
            })
        elif not is_medicaid and not is_recall_only and codes_set & COPAY_TRIGGER_CODES:
            alerts["copay"].append({
                "time": time_str, "patient": apt["PatientName"],
                "provider": apt["ProviderName"], "amount": "Check copay"
            })

        if has_ins and codes_set & PRIOR_AUTH_CODES:
            pa_codes = list(codes_set & PRIOR_AUTH_CODES)
            alerts["prior_auth"].append({
                "time": time_str, "patient": apt["PatientName"],
                "provider": apt["ProviderName"],
                "codes": ", ".join(pa_codes), "carrier": carrier
            })





    # -------------------------------------------------------------------------
    # 7. Schedule health per provider
    # -------------------------------------------------------------------------
    provider_health = {}

    # If no schedule table data, build from appointments
    if not prov_schedules and enriched:
        from collections import defaultdict
        apt_by_prov = defaultdict(list)
        for a in enriched:
            apt_by_prov[a["prov_num"]].append(a)
        for pn, apts in apt_by_prov.items():
            pname = apts[0]["provider"]
            times = sorted(a["time_min"] for a in apts)
            prov_schedules[pn] = {
                "name":      pname,
                "abbr":      apts[0]["prov_abbr"],
                "start_min": times[0],
                "stop_min":  times[-1] + 60,
                "start_str": "",
                "stop_str":  "",
            }

    for prov_num, sched in prov_schedules.items():
        pname = sched["name"]

        prov_apts = [a for a in enriched if a["prov_num"] == prov_num]
        if not prov_apts:
            continue

        total_sched_min = sched["stop_min"] - sched["start_min"]
        if total_sched_min <= 0:
            continue
        # Flag if this is a fallback schedule (no real schedule data)
        # Real schedule has start_str set from strftime; fallback sets it to ""
        is_fallback_sched = (sched.get("start_str", "") == "" and
                             sched.get("stop_str", "") == "")

        # Utilization (20 pts)
        used_min   = sum(a.get("duration_min", 30) for a in prov_apts)
        util_rate  = min(used_min / total_sched_min * 100, 100)
        util_score = 20 if util_rate >= 90 else 14 if util_rate >= 75 else 8 if util_rate >= 60 else 0

        # Confirmation rate (20 pts)
        CONF_OK_HEALTH = (
            "Appointment Confirmed", "Confirmed",
            "Arrived", "In Treatment Room", "In Room",
            "Ready", "Claim Sent", "No Bill to Insurance",
            "No Bill", "Complete"
        )
        confirmed   = sum(1 for a in prov_apts if a["confirmed"] in CONF_OK_HEALTH)
        conf_rate   = confirmed / len(prov_apts) * 100 if prov_apts else 0
        conf_score  = 20 if conf_rate >= 90 else 14 if conf_rate >= 75 else 8 if conf_rate >= 50 else 0

        # No-show risk (15 pts)
        high_risk   = sum(1 for a in prov_apts if a["noshow_risk"] == "high")
        risk_score  = 15 if high_risk == 0 else 9 if high_risk == 1 else 4 if high_risk == 2 else 0

        # Payer mix — PPO + No Insurance (25 pts)
        ppo_no_ins  = sum(1 for a in prov_apts
                         if not a["is_medicaid"] or not a["has_ins"])
        payer_pct   = ppo_no_ins / len(prov_apts) * 100 if prov_apts else 0
        payer_score = 25 if payer_pct >= 50 else 18 if payer_pct >= 35 else 10 if payer_pct >= 20 else 4 if payer_pct >= 10 else 0

        # Production value (10 pts)
        all_codes   = set(c for a in prov_apts for c in a["codes"])
        high_val    = len(all_codes & HIGH_VALUE_CODES)
        med_val     = len(all_codes & MEDIUM_VALUE_CODES)
        prod_score  = 10 if high_val >= 2 else 7 if high_val >= 1 else 4 if med_val >= 1 else 2

        # New patient mix (10 pts)
        new_pts     = sum(1 for a in prov_apts if a["is_new"])
        new_pct     = new_pts / len(prov_apts) * 100 if prov_apts else 0
        mix_score   = 10 if 10 <= new_pct <= 25 else 6 if 1 <= new_pct <= 40 else 3

        # Gaps for open chair time only (not scored)
        apt_times   = sorted([(a["time_min"], a["time_min"] + a.get("duration_min", 30))
                               for a in prov_apts])
        blocked     = [(b["start_min"], b["stop_min"]) for b in blockouts]
        busy        = sorted(apt_times + blocked)
        gaps        = 0
        prev_end    = sched["start_min"]
        for start, end in busy:
            if start > prev_end + 30:
                gaps += 1
            prev_end = max(prev_end, end)
        cont_score  = 0

        total_score = util_score + conf_score + risk_score + payer_score + prod_score + mix_score
        rating      = ("Strong" if total_score >= 85 else
                       "Good" if total_score >= 70 else
                       "At Risk" if total_score >= 50 else "Critical")
        color       = ("#2ecc71" if total_score >= 85 else
                       "#f1c40f" if total_score >= 70 else
                       "#e67e22" if total_score >= 50 else "#e74c3c")

        # Open chair slots
        open_slots  = []
        prev_end    = sched["start_min"]
        for start, end in sorted(busy):
            if start - prev_end >= 30:
                sh = prev_end // 60
                sm = prev_end % 60
                eh = start // 60
                em = start % 60
                ampm_s = "AM" if sh < 12 else "PM"
                ampm_e = "AM" if eh < 12 else "PM"
                sh12   = sh % 12 or 12
                eh12   = eh % 12 or 12
                open_slots.append(
                    f"{sh12}:{sm:02d}{ampm_s}–{eh12}:{em:02d}{ampm_e} "
                    f"({start - prev_end}min)"
                )
            prev_end = max(prev_end, end)
        # Gap after last appointment
        if sched["stop_min"] - prev_end >= 30:
            gap = sched["stop_min"] - prev_end
            sh  = prev_end // 60
            sm  = prev_end % 60
            eh  = sched["stop_min"] // 60
            em  = sched["stop_min"] % 60
            ampm_s = "AM" if sh < 12 else "PM"
            ampm_e = "AM" if eh < 12 else "PM"
            sh12   = sh % 12 or 12
            eh12   = eh % 12 or 12
            open_slots.append(
                f"{sh12}:{sm:02d}{ampm_s}–{eh12}:{em:02d}{ampm_e} ({gap}min)"
            )

        high_risk_patients = [a["patient"] for a in prov_apts
                               if a["noshow_risk"] == "high"]

        provider_health[prov_num] = {
            "name":         pname,
            "abbr":         sched["abbr"],
            "score":        total_score,
            "rating":       rating,
            "color":        color,
            "util_rate":    round(util_rate),
            "util_score":   util_score,
            "conf_rate":    round(conf_rate),
            "conf_score":   conf_score,
            "risk_score":   risk_score,
            "prod_score":   prod_score,
            "mix_score":    mix_score,
            "cont_score":   cont_score,
            "payer_score":  payer_score,
            "payer_pct":    round(payer_pct),
            "high_risk":    high_risk_patients,
            "open_slots":   open_slots,
            "total_apts":   len(prov_apts),
            "sched_hours":  f"{sched['start_str']} – {sched['stop_str']}",
            "has_high_val": high_val >= 1,
            "new_pts":      new_pts,
        }

    # -------------------------------------------------------------------------
    # 8. Provider schedule cards
    # -------------------------------------------------------------------------
    prov_cards = {}
    for prov_num, sched in prov_schedules.items():
        if sched["name"] == "Unknown":
            continue
        prov_apts = [a for a in enriched if a["prov_num"] == prov_num]
        prov_cards[prov_num] = {
            "name":     sched["name"],
            "abbr":     sched["abbr"],
            "hours":    f"{sched['start_str']} – {sched['stop_str']}",
            "apts":     prov_apts,
            "health":   provider_health.get(prov_num) or provider_health.get(str(prov_num)),
        }

    # If schedule table returned nothing, build provider cards from appointments
    if not prov_cards and enriched:
        seen_provs = {}
        for apt in enriched:
            pn = apt["prov_num"]
            if pn not in seen_provs:
                seen_provs[pn] = {
                    "name":   apt["provider"],
                    "abbr":   apt["prov_abbr"],
                    "hours":  "See schedule",
                    "apts":   [],
                    "health": provider_health.get(pn),
                }
            seen_provs[pn]["apts"].append(apt)
        prov_cards = seen_provs
    else:
        # For providers with schedule but excluded from health,
        # still build their card
        for apt in enriched:
            pn = apt["prov_num"]
            pname = apt["provider"]
    



    # -------------------------------------------------------------------------
    # DAY SUMMARY — cancelled, no-shows, rescheduled, production
    # -------------------------------------------------------------------------
    # Cancelled / no-shows today (AptStatus 5 = Broken, 3 = UnschedList)
    broken_today = [
        a for a in enriched if a["apt_status"] in (3, 5)
    ]
    active_today = [
        a for a in enriched if a["apt_status"] not in (3, 5)
    ]
    completed_today = [
        a for a in enriched if a["apt_status"] == 2
    ]
    still_scheduled = [
        a for a in enriched if a["apt_status"] == 1
    ]

    # Check rescheduled: two cases
    # 1. Broken/cancelled today → new future appointment booked
    # 2. Appointment originally on today → moved to a future date (histappointment)
    rescheduled_pat_nums = set()

    # Case 1: broken/cancelled patients with a future appointment
    if broken_today:
        broken_pat_nums = [a["pat_num"] for a in broken_today]
        fmt = ",".join(["%s"] * len(broken_pat_nums))
        cursor.execute(f"""
            SELECT DISTINCT PatNum
            FROM appointment
            WHERE PatNum IN ({fmt})
            AND AptStatus IN (1, 2)
            AND DATE(AptDateTime) > %s
        """, broken_pat_nums + [snap_date])
        for row in cursor.fetchall():
            rescheduled_pat_nums.add(row["PatNum"])

    # Case 2: appointments originally scheduled today that were moved to a future date
    # histappointment logs every change — find apts whose old date was today
    # but current AptDateTime is a future date (i.e. they were moved, not cancelled)
    try:
        cursor.execute("""
            SELECT DISTINCT a.PatNum
            FROM histappointment h
            INNER JOIN appointment a ON a.AptNum = h.AptNum
            WHERE DATE(h.AptDateTime) = %s
            AND h.HistApptAction = 1
            AND DATE(a.AptDateTime) > %s
            AND a.AptStatus IN (1, 2)
            AND DATE(h.HistDateTStamp) = %s
        """, (snap_date, snap_date, snap_date))
        for row in cursor.fetchall():
            rescheduled_pat_nums.add(row["PatNum"])
    except Exception:
        pass  # histappointment may not exist in all versions

    # Case 3: mid-day reschedules — appointments that were on today's schedule
    # but were moved to a future date during the day (not cancelled, just rescheduled).
    # These disappear from today's query entirely so we detect them via histappointment.
    # Note: HistApptAction is unreliable (0 or 1 for same operation), so we filter
    # purely on the appointment having moved to a future scheduled date.
    midday_rescheduled = []
    today_apt_nums  = set(a["aptnum"] for a in enriched)
    broken_apt_nums = set(a["aptnum"] for a in broken_today)
    try:
        cursor.execute("""
            SELECT DISTINCT h.AptNum,
                   DATE_FORMAT(h.AptDateTime, '%h:%i %p') AS OldTime,
                   CONCAT(pat.LName, ', ', pat.FName) AS PatientName,
                   COALESCE(NULLIF(pat.WirelessPhone,''),
                            NULLIF(pat.HmPhone,''),
                            NULLIF(pat.WkPhone,'')) AS Phone,
                   CONCAT(prov.FName, ' ', prov.LName) AS ProviderName,
                   DATE_FORMAT(a.AptDateTime, '%m/%d/%Y') AS NewDate
            FROM histappointment h
            INNER JOIN appointment a ON a.AptNum = h.AptNum
            INNER JOIN patient pat ON pat.PatNum = a.PatNum
            INNER JOIN provider prov ON prov.ProvNum = a.ProvNum
            WHERE DATE(h.AptDateTime) = %s
            AND DATE(a.AptDateTime) > %s
            AND a.AptStatus = 1
            AND DATE(h.HistDateTStamp) = %s
        """, (snap_date, snap_date, snap_date))
        for row in cursor.fetchall():
            if row["AptNum"] in today_apt_nums:
                continue
            if row["AptNum"] in broken_apt_nums:
                continue
            midday_rescheduled.append({
                "time":     (row["OldTime"] or "").lstrip("0"),
                "patient":  row["PatientName"],
                "provider": row["ProviderName"],
                "phone":    row["Phone"] or "",
                "new_date": row["NewDate"] or "",
            })
    except Exception:
        pass

    # Case 4: deleted appointments — were on today's schedule but deleted entirely.
    # HistApptAction=4, appointment no longer exists in appointment table.
    deleted_today = []
    try:
        cursor.execute("""
            SELECT DISTINCT h.AptNum, h.PatNum,
                   DATE_FORMAT(h.AptDateTime, '%h:%i %p') AS OldTime,
                   CONCAT(pat.LName, ', ', pat.FName) AS PatientName,
                   COALESCE(NULLIF(pat.WirelessPhone,''),
                            NULLIF(pat.HmPhone,''),
                            NULLIF(pat.WkPhone,'')) AS Phone,
                   CONCAT(prov.FName, ' ', prov.LName) AS ProviderName
            FROM histappointment h
            LEFT JOIN appointment a ON a.AptNum = h.AptNum
            INNER JOIN patient pat ON pat.PatNum = h.PatNum
            INNER JOIN provider prov ON prov.ProvNum = h.ProvNum
            WHERE DATE(h.AptDateTime) = %s
            AND h.HistApptAction = 4
            AND a.AptNum IS NULL
            AND DATE(h.HistDateTStamp) = %s
        """, (snap_date, snap_date))
        for row in cursor.fetchall():
            deleted_today.append({
                "time":     (row["OldTime"] or "").lstrip("0"),
                "patient":  row["PatientName"],
                "provider": row["ProviderName"],
                "phone":    row["Phone"] or "",
            })
    except Exception:
        pass

    # Day production (completed procedures today)
    cursor.execute("""
        SELECT COALESCE(SUM(pl.ProcFee), 0) AS DayProd,
               COUNT(*) AS ProcCount
        FROM procedurelog pl
        WHERE pl.ProcStatus = 2
        AND pl.ProcDate = %s
    """, (snap_date,))
    prod_row  = cursor.fetchone()
    day_prod  = float(prod_row["DayProd"] or 0)
    day_procs = int(prod_row["ProcCount"] or 0)

    # Day mode: summary when office is done for the day (today only, never for future dates)
    from datetime import datetime as _dt
    now_hour = _dt.now().hour
    is_future = snap_date > date.today()
    day_mode = (
        "summary" if (not is_future and len(still_scheduled) == 0 and len(completed_today) > 0 and now_hour >= 15)
        else "active"
    )

    # Build cancelled list with rescheduled flag
    cancelled_list = []
    for a in broken_today:
        cancelled_list.append({
            "time":        a["time"],
            "patient":     a["patient"],
            "provider":    a["provider"],
            "phone":       a["phone"],
            "apt_status":  a["apt_status"],
            "label":       "No-Show" if a["apt_status"] == 5 else "Unscheduled",
            "rescheduled": a["pat_num"] in rescheduled_pat_nums,
        })

    # Per-provider day summary
    prov_day = {}
    for a in enriched:
        pn = a["prov_num"]
        if pn not in prov_day:
            prov_day[pn] = {
                "name": a["provider"], "abbr": a["prov_abbr"],
                "seen": 0, "cancelled": 0, "scheduled": 0,
            }
        s = a["apt_status"]
        if s == 2:
            prov_day[pn]["seen"] += 1
        elif s in (3, 5):
            prov_day[pn]["cancelled"] += 1
        else:
            prov_day[pn]["scheduled"] += 1

    # Build missing routing slips list from completed appointments
    missing_slips = [
        {"time": a["time"], "patient": a["patient"], "provider": a["provider"]}
        for a in completed_today
        if not a.get("routing_slip", False)
    ]

    day_summary = {
        "mode":               day_mode,
        "completed":          len(completed_today),
        "still_active":       len(still_scheduled),
        "cancelled":          len(broken_today),
        "cancelled_list":     cancelled_list,
        "rescheduled":        len(midday_rescheduled),
        "rescheduled_list":   midday_rescheduled,
        "deleted":            len(deleted_today),
        "deleted_list":       deleted_today,
        "missing_slips":      len(missing_slips),
        "missing_slips_list": missing_slips,
        "prov_day":           list(prov_day.values()),
        "day_prod":           round(day_prod),
        "day_procs":          day_procs,
    }

    # Convert int keys to strings for JSON
    prov_cards_str = {str(k): v for k, v in prov_cards.items()}

    return {
        "date":        today_str,
        "snap_date":   snap_date_str,
        "is_preview":  is_preview,
        "total_pts":   len([a for a in enriched if a["apt_status"] not in (3, 5)]),
        "alerts":      alerts,
        "prov_cards":  prov_cards_str,
        "provider_health": {str(k): v for k, v in provider_health.items()},
        "blockouts":   blockouts,
        "day_summary": day_summary,
    }


# =============================================================================
# FLASK
# =============================================================================

app = Flask(__name__)

# ── Auth setup ────────────────────────────────────────────────────────────────
import functools, hashlib

# Stable secret derived from practice name — sessions survive dashboard restart
app.secret_key = hashlib.sha256(f"dental-{PRACTICE_NAME}".encode()).hexdigest()
# Session cookies are non-permanent by default — they expire when the browser closes.

def login_required(view):
    @functools.wraps(view)
    def wrapper(*args, **kwargs):
        if DASHBOARD_PASSWORD and not session.get("authed"):
            # API/POST → 401 JSON; GETs → redirect to login page
            if request.method != "GET" or request.path.startswith("/api/"):
                return jsonify({"error": "auth required"}), 401
            return redirect("/login")
        return view(*args, **kwargs)
    return wrapper

LOGIN_HTML = r"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"><title>__PRACTICE__</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@400;500;700&display=swap" rel="stylesheet">
<style>
  *{box-sizing:border-box;margin:0;padding:0}
  body{
    min-height:100vh;
    background:#070708;
    background-image:
      linear-gradient(rgba(34,211,238,0.025) 1px, transparent 1px),
      linear-gradient(90deg, rgba(34,211,238,0.025) 1px, transparent 1px);
    background-size:44px 44px;
    color:#e2e8f0;
    font-family:'JetBrains Mono',ui-monospace,monospace;
    display:flex;align-items:center;justify-content:center;padding:24px;
  }
  .card{
    position:relative;width:400px;padding:36px 32px;
    background:rgba(0,0,0,0.55);
    border:1px solid rgba(34,211,238,0.35);
    box-shadow:0 0 60px rgba(34,211,238,0.06),inset 0 0 0 1px rgba(34,211,238,0.04);
  }
  .card::before,.card::after{
    content:'';position:absolute;width:14px;height:14px;
    border:1px solid #22d3ee;pointer-events:none;
  }
  .card::before{top:-1px;left:-1px;border-right:none;border-bottom:none}
  .card::after{bottom:-1px;right:-1px;border-left:none;border-top:none}

  .tag{
    display:flex;align-items:center;gap:8px;
    color:#22d3ee;font-size:10px;letter-spacing:3px;font-weight:700;
    margin-bottom:14px;
  }
  .tag::before{
    content:'';width:6px;height:6px;background:#22d3ee;
    box-shadow:0 0 10px #22d3ee;
  }
  .heading{
    font-size:22px;font-weight:700;color:#f0f9ff;
    letter-spacing:1px;line-height:1.2;margin-bottom:6px;
  }
  .sub{
    font-size:11px;color:#475569;letter-spacing:2px;
    margin-bottom:32px;
  }

  .field-label{
    font-size:10px;color:#64748b;text-transform:uppercase;
    letter-spacing:3px;margin-bottom:8px;
  }
  .input-wrap{
    display:flex;align-items:center;
    background:#000;border:1px solid rgba(34,211,238,0.3);
    transition:border-color .15s,box-shadow .15s;
  }
  .input-wrap:focus-within{
    border-color:#22d3ee;
    box-shadow:0 0 0 1px rgba(34,211,238,0.2),0 0 24px rgba(34,211,238,0.08);
  }
  .input-wrap::before{
    content:'>';color:#22d3ee;padding:0 14px;font-weight:700;font-size:14px;
  }
  input{
    flex:1;padding:14px 14px 14px 0;background:transparent;border:none;
    color:#f0f9ff;font-family:inherit;font-size:14px;letter-spacing:3px;outline:none;
  }
  input::placeholder{color:#1e293b;letter-spacing:0}

  button{
    width:100%;margin-top:24px;padding:14px;background:transparent;
    border:1px solid #22d3ee;color:#22d3ee;font-family:inherit;
    font-size:11px;font-weight:700;letter-spacing:4px;text-transform:uppercase;
    cursor:pointer;transition:all .15s;
  }
  button:hover{
    background:#22d3ee;color:#000;
    box-shadow:0 0 24px rgba(34,211,238,0.4);
  }
  button:active{transform:translateY(1px)}

  .err{
    margin-top:16px;padding:10px 12px;
    background:rgba(220,38,38,0.08);border:1px solid rgba(220,38,38,0.4);
    color:#fca5a5;font-size:11px;letter-spacing:1.5px;
  }
  .err::before{content:'[ ERR ] ';color:#dc2626;font-weight:700}

  .footer{
    margin-top:24px;font-size:9px;color:#334155;
    text-align:center;letter-spacing:3px;
  }
</style></head><body>
  <form class="card" method="post" action="/login">
    <div class="tag">SECURE ACCESS</div>
    <div class="heading">__PRACTICE__</div>
    <div class="sub">// AUTH REQUIRED</div>
    <div class="field-label">[ PASSWORD ]</div>
    <div class="input-wrap">
      <input type="password" name="password" autofocus required autocomplete="current-password">
    </div>
    <button>AUTHENTICATE</button>
    __ERROR__
    <div class="footer">// CONTROL PANEL // LOCAL_NET</div>
  </form>
</body></html>"""

@app.route("/login", methods=["GET", "POST"])
def login():
    if not DASHBOARD_PASSWORD:
        return redirect("/")
    if session.get("authed"):
        return redirect("/")
    error_html = ""
    if request.method == "POST":
        pw = (request.form.get("password") or "").strip()
        if pw == DASHBOARD_PASSWORD:
            session["authed"] = True
            return redirect("/")
        error_html = '<div class="err">ACCESS DENIED // RETRY</div>'
    page = LOGIN_HTML.replace("__PRACTICE__", PRACTICE_NAME).replace("__ERROR__", error_html)
    return page

@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")

HTML = r"""<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<title>__PRACTICE__ - Control Panel</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&display=swap" rel="stylesheet">
<style>
/* -----------------------------------------------
   PRISTINE DENTAL  REDESIGNED DASHBOARD STYLES
   Clean, modern, legible. Inter typeface.
----------------------------------------------- */

* { margin:0; padding:0; box-sizing:border-box }

body {
  font-family: 'Inter', 'Segoe UI', system-ui, sans-serif;
  background: #f4f6fb;
  color: #0f172a;
  min-height: 100vh;
  font-size: 14px;
  line-height: 1.5;
  -webkit-font-smoothing: antialiased;
}

/* -- HEADER -- */
.hdr {
  background: linear-gradient(135deg, #0f2250 0%, #1a3a7c 60%, #1e4db7 100%);
  padding: 14px 32px;
  display: flex;
  align-items: center;
  justify-content: space-between;
  position: sticky;
  top: 0;
  z-index: 100;
  box-shadow: 0 1px 0 rgba(255,255,255,0.08), 0 4px 20px rgba(10,30,90,.35);
}
.hdr h1 {
  font-size: 20px;
  font-weight: 800;
  color: #ffffff;
  letter-spacing: -0.5px;
}
.hdr p {
  font-size: 12px;
  color: rgba(147, 197, 253, 0.85);
  font-weight: 500;
  letter-spacing: 0.3px;
  margin-top: 1px;
}
.clk {
  font-size: 16px;
  font-weight: 700;
  background: rgba(255,255,255,0.1);
  color: #fde68a;
  padding: 7px 18px;
  border-radius: 10px;
  border: 1px solid rgba(253,230,138,0.25);
  letter-spacing: 0.5px;
}

/* -- CONTROL BAR -- */
.ctrl {
  padding: 10px 32px;
  background: #ffffff;
  border-bottom: 1px solid #e8edf5;
  display: flex;
  gap: 8px;
  box-shadow: 0 1px 4px rgba(0,0,0,.04);
  align-items: center;
}

/* -- BUTTONS -- */
.btn {
  padding: 8px 18px;
  border: none;
  border-radius: 8px;
  font-size: 13px;
  font-weight: 600;
  cursor: pointer;
  transition: all 0.15s ease;
  font-family: inherit;
  letter-spacing: 0.1px;
}
.btn:hover { filter: brightness(1.1); transform: translateY(-1px); box-shadow: 0 3px 8px rgba(0,0,0,.12); }
.btn:active { transform: translateY(0); }
.bg { background: #16a34a; color: #fff; }
.br { background: #dc2626; color: #fff; }
.bb { background: #2563eb; color: #fff; }
.bo { background: #d97706; color: #fff; }
.btn:disabled { opacity: .4; cursor: not-allowed; transform: none !important; box-shadow: none !important; }

.wrap {
  display: grid;
  grid-template-columns: 1fr 340px;
  gap: 20px;
  padding: 20px 32px;
  max-width: 1440px;
  margin: 0 auto;
}
.scripts { display: flex; flex-direction: column; gap: 14px; }

/* -- SCRIPT CARDS -- */
.card {
  background: #ffffff;
  border: 1px solid #e8edf5;
  border-radius: 14px;
  overflow: hidden;
  box-shadow: 0 1px 3px rgba(0,0,0,.05), 0 4px 12px rgba(0,0,0,.04);
  transition: box-shadow 0.2s;
}
.card:hover { box-shadow: 0 2px 8px rgba(0,0,0,.08), 0 8px 24px rgba(0,0,0,.06); }
.card.paused { opacity: .7; }
.card.running { border-color: #f59e0b; box-shadow: 0 0 0 2px rgba(245,158,11,.15); }
.ch {
  padding: 14px 18px;
  display: flex;
  align-items: center;
  justify-content: space-between;
  border-bottom: 1px solid #f1f5f9;
  background: linear-gradient(135deg, #f8faff, #f1f5ff);
}
.ct { display: flex; align-items: center; gap: 10px; }
.ci {
  font-size: 12px;
  font-weight: 700;
  width: 88px;
  padding: 5px 8px;
  text-align: center;
  border-radius: 8px;
  background: #eff6ff;
  color: #1d4ed8;
  border: 1px solid #bfdbfe;
  letter-spacing: 0.2px;
}
.cn { font-size: 16px; font-weight: 700; color: #0f172a; }
.cd { font-size: 12px; color: #64748b; margin-top: 2px; }
.cbtns { display: flex; gap: 6px; align-items: center; }
.badge {
  font-size: 11px;
  padding: 3px 10px;
  border-radius: 20px;
  font-weight: 700;
  letter-spacing: 0.3px;
}
.ba { background: #dcfce7; color: #15803d; border: 1px solid #86efac; }
.bp { background: #f1f5f9; color: #94a3b8; border: 1px solid #e2e8f0; }
.brun { background: #fef3c7; color: #b45309; border: 1px solid #fde68a; }
.bsm { padding: 5px 12px; font-size: 12px; }
.cb { padding: 14px 18px; display: grid; grid-template-columns: 1fr 1fr; gap: 10px; }
.info label {
  font-size: 10px;
  color: #94a3b8;
  text-transform: uppercase;
  letter-spacing: 0.8px;
  font-weight: 600;
}
.info p { font-size: 13px; margin-top: 2px; font-weight: 600; color: #0f172a; }
.sum {
  grid-column: 1/-1;
  background: #f8fafc;
  border: 1px solid #e8edf5;
  border-radius: 8px;
  padding: 9px 12px;
  font-size: 12px;
  color: #64748b;
  line-height: 1.6;
}
.ok { color: #16a34a; }
.er { color: #dc2626; }
.spin {
  display: inline-block; width: 12px; height: 12px;
  border: 2px solid rgba(217,119,6,.25); border-top-color: #d97706;
  border-radius: 50%; animation: sp .8s linear infinite;
}
@keyframes sp { to { transform: rotate(360deg) } }

/* -- APPOINTMENT STATUS -- */
.apt-row.apt-complete { opacity: .6; background: rgba(22,163,74,.03); }
.apt-row.apt-complete .apt-patient { text-decoration: line-through; color: #94a3b8; }
.apt-row.apt-broken { opacity: .5; background: rgba(220,38,38,.03); }
.apt-row.apt-broken .apt-patient { text-decoration: line-through; color: #94a3b8; }
.apt-status-dot { width: 8px; height: 8px; border-radius: 50%; flex-shrink: 0; margin-top: 6px; }
.dot-scheduled  { background: #3b82f6; }
.dot-complete   { background: #16a34a; }
.dot-broken     { background: #dc2626; }
.dot-unscheduled { background: #d97706; }

/* -- DAY SUMMARY -- */
.day-summary-wrap {
  background: #f4f6fb;
  padding: 0 32px 20px;
  border-bottom: 1px solid #e8edf5;
}
.day-summary-banner {
  background: linear-gradient(135deg, #fffbeb, #fef9ec);
  border: 1px solid #fde68a;
  border-radius: 14px;
  padding: 16px 24px;
  margin-bottom: 14px;
  display: flex;
  align-items: center;
  justify-content: space-between;
  box-shadow: 0 2px 12px rgba(217,119,6,.08);
}
.day-summary-banner h3 { color: #78350f; font-size: 18px; font-weight: 800; margin: 0; letter-spacing: -0.3px; }
.day-summary-banner p  { color: #92400e; font-size: 12px; margin: 3px 0 0; font-weight: 500; }
.day-stats-row { display: flex; gap: 10px; flex-wrap: wrap; margin-bottom: 14px; }
.day-stat {
  background: #ffffff;
  border: 1px solid #e8edf5;
  border-radius: 12px;
  flex: 1;
  min-width: 110px;
  padding: 16px 18px;
  text-align: center;
  box-shadow: 0 1px 4px rgba(0,0,0,.04);
  transition: transform 0.15s;
}
.day-stat:hover { transform: translateY(-1px); box-shadow: 0 4px 12px rgba(0,0,0,.07); }
.day-stat-num { font-size: 30px; font-weight: 800; line-height: 1; letter-spacing: -1px; }
.day-stat-lbl { font-size: 11px; color: #64748b; text-transform: uppercase; letter-spacing: 0.7px; font-weight: 600; margin-top: 5px; }
.day-prov-grid { display: flex; gap: 10px; flex-wrap: wrap; margin-bottom: 14px; }
.day-prov-card {
  background: #ffffff;
  border: 1px solid #e8edf5;
  border-radius: 12px;
  flex: 1;
  min-width: 160px;
  padding: 14px 16px;
  box-shadow: 0 1px 4px rgba(0,0,0,.04);
}
.day-prov-name { font-size: 14px; font-weight: 700; color: #0f172a; margin-bottom: 8px; }
.day-prov-row {
  display: flex;
  justify-content: space-between;
  font-size: 12px;
  color: #64748b;
  padding: 4px 0;
  border-bottom: 1px solid #f1f5f9;
}
.day-prov-row:last-child { border-bottom: none; }
.day-prov-val { font-weight: 700; color: #1e3a8a; }
.cancelled-list {
  background: #ffffff;
  border: 1px solid #e8edf5;
  border-radius: 12px;
  padding: 16px 20px;
  box-shadow: 0 1px 4px rgba(0,0,0,.04);
}
.cancelled-hdr {
  font-size: 11px;
  color: #94a3b8;
  font-weight: 700;
  text-transform: uppercase;
  letter-spacing: 0.9px;
  margin-bottom: 12px;
  padding-bottom: 8px;
  border-bottom: 1px solid #f1f5f9;
}
.cancelled-item {
  display: flex;
  align-items: center;
  justify-content: space-between;
  padding: 10px 0;
  border-bottom: 1px solid #f8fafc;
  gap: 12px;
}
.cancelled-item:last-child { border-bottom: none; }
.cancelled-name { font-size: 14px; font-weight: 600; color: #0f172a; }
.cancelled-meta { font-size: 12px; color: #64748b; margin-top: 2px; }
.resched-badge {
  font-size: 11px; font-weight: 700; padding: 3px 10px; border-radius: 20px;
  background: #dcfce7; color: #15803d; border: 1px solid #86efac; white-space: nowrap;
}
.noshow-badge {
  font-size: 11px; font-weight: 700; padding: 3px 10px; border-radius: 20px;
  background: #fee2e2; color: #dc2626; border: 1px solid #fca5a5; white-space: nowrap;
}
.unsched-badge {
  font-size: 11px; font-weight: 700; padding: 3px 10px; border-radius: 20px;
  background: #fef3c7; color: #b45309; border: 1px solid #fde68a; white-space: nowrap;
}

/* -- SIDE PANEL -- */
.lpanel {
  background: #ffffff;
  border: 1px solid #e8edf5;
  border-radius: 14px;
  overflow: hidden;
  position: sticky;
  top: 60px;
  box-shadow: 0 2px 8px rgba(0,0,0,.06);
}
.lh {
  padding: 12px 16px;
  border-bottom: 1px solid #e8edf5;
  font-weight: 700;
  font-size: 13px;
  color: #1e3a8a;
  background: linear-gradient(135deg, #f0f4ff, #f8faff);
  letter-spacing: 0.1px;
}
.lb { max-height: 72vh; overflow-y: auto; }
.li { padding: 10px 16px; border-bottom: 1px solid #f1f5f9; font-size: 13px; }
.lt { color: #94a3b8; font-size: 11px; font-weight: 500; }
.ls { font-weight: 600; margin: 2px 0; color: #0f172a; }
.lem { padding: 28px 16px; text-align: center; color: #94a3b8; font-size: 13px; }

/* -- SNAPSHOT HEADER -- */
.snap-hdr {
  background: linear-gradient(135deg, #0f2250, #1a3a7c);
  padding: 18px 32px;
  display: flex;
  align-items: center;
  justify-content: space-between;
  border-bottom: none;
  box-shadow: 0 2px 8px rgba(10,30,90,.25);
}
.snap-hdr h2 { color: #ffffff; font-size: 19px; font-weight: 800; margin: 0; letter-spacing: -0.4px; }
.preview-badge {
  background: #f59e0b; color: #fff; padding: 4px 12px; border-radius: 20px;
  font-size: 11px; font-weight: 700; margin-left: 10px; letter-spacing: 0.5px;
}

/* -- STATS BAR -- */
.stats-bar {
  display: flex;
  background: #ffffff;
  border-bottom: 2px solid #e8edf5;
}
.stat-box {
  flex: 1;
  padding: 18px 24px;
  border-right: 1px solid #f1f5f9;
  transition: background 0.15s;
  position: relative;
  cursor: default;
}
.stat-box:last-child { border-right: none; }
.stat-box:hover { background: #f8faff; }
.stat-num {
  font-size: 36px;
  font-weight: 800;
  line-height: 1;
  letter-spacing: -1.5px;
  color: #1e3a8a;
}
.stat-lbl {
  font-size: 11px;
  color: #94a3b8;
  text-transform: uppercase;
  letter-spacing: 0.8px;
  margin-top: 6px;
  font-weight: 600;
}
.stat-box.danger .stat-num { color: #dc2626; }
.stat-box.warn   .stat-num { color: #d97706; }
.stat-box.ok     .stat-num { color: #16a34a; }
.stat-box.neutral .stat-num { color: #64748b; }

/* -- PROVIDER ROW -- */
.prov-row {
  display: flex;
  gap: 16px;
  padding: 20px 32px;
  overflow-x: auto;
  background: #f4f6fb;
  border-bottom: 1px solid #e8edf5;
}
.prov-row::-webkit-scrollbar { height: 4px; }
.prov-row::-webkit-scrollbar-track { background: #e8edf5; }
.prov-row::-webkit-scrollbar-thumb { background: #93c5fd; border-radius: 2px; }
.prov-card {
  background: #ffffff;
  border: 1px solid #e8edf5;
  border-radius: 14px;
  min-width: 290px;
  max-width: 330px;
  flex-shrink: 0;
  overflow: hidden;
  transition: all 0.2s;
  box-shadow: 0 1px 4px rgba(0,0,0,.05), 0 4px 12px rgba(15,34,80,.05);
}
.prov-card:hover {
  border-color: #93c5fd;
  box-shadow: 0 4px 20px rgba(15,34,80,.12);
  transform: translateY(-2px);
}
.prov-card-hdr {
  padding: 14px 16px;
  border-bottom: 1px solid #e8edf5;
  background: linear-gradient(135deg, #0f2250, #1a3a7c);
}
.prov-name { font-size: 15px; font-weight: 700; color: #ffffff; letter-spacing: -0.2px; }
.prov-meta { font-size: 11px; color: #93c5fd; margin-top: 3px; font-weight: 500; }

/* -- APPOINTMENT LIST -- */
.apt-list { max-height: 240px; overflow-y: auto; }
.apt-list::-webkit-scrollbar { width: 3px; }
.apt-list::-webkit-scrollbar-thumb { background: #bfdbfe; border-radius: 2px; }
.apt-row {
  padding: 9px 16px;
  border-bottom: 1px solid #f8fafc;
  display: flex;
  align-items: flex-start;
  gap: 10px;
  transition: background 0.12s;
}
.apt-row:hover { background: #f8faff; }
.apt-row:last-child { border-bottom: none; }
.apt-time {
  color: #1d4ed8;
  font-weight: 700;
  font-size: 13px;
  min-width: 58px;
  white-space: nowrap;
  padding-top: 1px;
}
.apt-info { flex: 1; min-width: 0; }
.apt-patient {
  color: #0f172a;
  font-weight: 600;
  font-size: 14px;
  white-space: nowrap;
  overflow: hidden;
  text-overflow: ellipsis;
}
.apt-meta {
  color: #64748b;
  font-size: 11px;
  margin-top: 2px;
  white-space: nowrap;
  overflow: hidden;
  text-overflow: ellipsis;
}
.apt-flags { display: flex; gap: 4px; flex-wrap: wrap; margin-top: 4px; }
.flag {
  padding: 2px 8px;
  border-radius: 20px;
  font-size: 10px;
  font-weight: 700;
  letter-spacing: 0.3px;
}
.flag-green { background: #dcfce7; color: #15803d; border: 1px solid #86efac; }
.flag-blue  { background: #dbeafe; color: #1d4ed8; border: 1px solid #93c5fd; }
.flag-red   { background: #fee2e2; color: #dc2626; border: 1px solid #fca5a5; }
.flag-yellow { background: #fef3c7; color: #b45309; border: 1px solid #fde68a; }

/* -- SCHEDULE HEALTH -- */
.health-section {
  padding: 14px 16px;
  background: #fafbff;
  border-top: 1px solid #e8edf5;
}
.health-top { display: flex; justify-content: space-between; align-items: center; margin-bottom: 6px; }
.health-label { font-size: 10px; color: #94a3b8; text-transform: uppercase; letter-spacing: 0.8px; font-weight: 600; }
.health-score { font-size: 22px; font-weight: 800; line-height: 1; }
.health-bar { height: 6px; border-radius: 3px; background: #e2e8f0; margin-bottom: 8px; overflow: hidden; }
.health-fill { height: 100%; border-radius: 3px; transition: width .6s ease; }
.health-rating { font-size: 12px; font-weight: 700; margin-bottom: 8px; }
.health-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 3px; }
.hd-item { display: flex; justify-content: space-between; font-size: 12px; color: #64748b; padding: 3px 0; }
.hd-val { color: #1e3a8a; font-weight: 600; }
.hd-val.good { color: #16a34a; }
.hd-val.bad  { color: #dc2626; }
.open-slots { margin-top: 8px; padding-top: 8px; border-top: 1px solid #e8edf5; }
.open-lbl { font-size: 10px; color: #94a3b8; text-transform: uppercase; letter-spacing: 0.8px; font-weight: 600; margin-bottom: 4px; }
.open-slot { font-size: 12px; color: #16a34a; padding: 2px 0; font-weight: 500; }

/* -- ALERT GRID -- */
.alert-grid {
  display: grid;
  grid-template-columns: repeat(4, 1fr);
  gap: 14px;
  padding: 20px 32px;
  background: #f4f6fb;
  border-bottom: 1px solid #e8edf5;
}
.alert-col {
  background: #ffffff;
  border-radius: 12px;
  overflow: hidden;
  box-shadow: 0 1px 4px rgba(15,34,80,.06), 0 4px 12px rgba(15,34,80,.04);
  border: 1px solid #e8edf5;
  transition: box-shadow 0.2s;
}
.alert-col:hover { box-shadow: 0 2px 8px rgba(15,34,80,.1), 0 8px 24px rgba(15,34,80,.07); }
.alert-col-hdr {
  padding: 11px 16px;
  font-size: 12px;
  font-weight: 700;
  letter-spacing: 0.3px;
  border-left: 4px solid transparent;
}
.alert-col-body { max-height: 300px; overflow-y: auto; }
.alert-col-body::-webkit-scrollbar { width: 3px; }
.alert-col-body::-webkit-scrollbar-thumb { background: #bfdbfe; border-radius: 2px; }
.alert-item {
  padding: 10px 14px;
  border-bottom: 1px solid #f8fafc;
  border-left: 4px solid transparent;
  transition: background 0.12s;
}
.alert-item:hover { background: #f8faff; }
.alert-item:last-child { border-bottom: none; }
.alert-pt { font-weight: 600; color: #0f172a; font-size: 13px; }
.alert-time { color: #1d4ed8; font-size: 11px; margin-bottom: 2px; font-weight: 600; }
.alert-sub { color: #64748b; font-size: 11px; margin-top: 3px; line-height: 1.5; }
.alert-empty { padding: 28px 16px; text-align: center; color: #cbd5e1; font-size: 13px; }

/* -- SCRIPTS SECTION -- */
.scripts-section { padding: 18px 32px 0; border-top: 1px solid #e8edf5; }
.scripts-lbl {
  font-size: 11px;
  color: #94a3b8;
  font-weight: 700;
  text-transform: uppercase;
  letter-spacing: 0.9px;
  margin-bottom: 12px;
}

/* AR aging buckets */
.ar-buckets { display: grid; grid-template-columns: repeat(4,1fr); gap: 12px; margin-bottom: 20px; }
.ar-bucket {
  background: #fff;
  border: 1px solid #e8edf5;
  border-radius: 12px;
  padding: 16px 18px;
  box-shadow: 0 1px 4px rgba(0,0,0,.04);
  transition: transform 0.15s;
}
.ar-bucket:hover { transform: translateY(-1px); box-shadow: 0 4px 14px rgba(0,0,0,.07); }
.ar-bucket-label { font-size: 11px; font-weight: 600; color: #94a3b8; text-transform: uppercase; letter-spacing: 0.7px; }
.ar-bucket-amount { font-size: 26px; font-weight: 800; margin: 4px 0 2px; letter-spacing: -0.8px; }
.ar-bucket-claims { font-size: 12px; color: #94a3b8; font-weight: 500; }

/* Age badges */
.age-0-30  { background: #dcfce7; color: #15803d; border: 1px solid #86efac; }
.age-31-60 { background: #fef3c7; color: #b45309; border: 1px solid #fde68a; }
.age-61-90 { background: #ffedd5; color: #c2410c; border: 1px solid #fed7aa; }
.age-90p   { background: #fee2e2; color: #dc2626; border: 1px solid #fca5a5; }

/* Scrollbar global */
::-webkit-scrollbar { width: 5px; height: 5px; }
::-webkit-scrollbar-track { background: #f1f5f9; }
::-webkit-scrollbar-thumb { background: #cbd5e1; border-radius: 3px; }
::-webkit-scrollbar-thumb:hover { background: #94a3b8; }

/* Insurance search */
.ins-search-wrap { padding: 20px 32px; background: #f4f6fb; border-top: 1px solid #e8edf5; }
.ins-search-inner { background: #fff; border-radius: 12px; padding: 16px 20px; box-shadow: 0 1px 4px rgba(0,0,0,.04); border: 1px solid #e8edf5; }
.ins-search-title { font-size: 13px; font-weight: 700; color: #0f172a; margin-bottom: 10px; }
.ins-search-row { display: flex; gap: 8px; }

/* Misc utility */
.text-muted { color: #64748b; }
.text-green { color: #16a34a; }
.text-red   { color: #dc2626; }
.text-blue  { color: #1d4ed8; }
.text-amber { color: #d97706; }
</style>
</head>
<body>
<div class="hdr">
  <div><h1>__PRACTICE__</h1><p>Automation Control Panel</p></div>
  <div style="display:flex;align-items:center;gap:12px">
    <a href="http://localhost:5002" target="_blank"
       style="padding:7px 16px;background:rgba(253,230,138,.15);color:#fde68a;
              border:1px solid rgba(253,230,138,.3);border-radius:8px;
              font-size:13px;font-weight:700;text-decoration:none">
      📊 AR Dashboard
    </a>
    <div class="clk" id="clk">--:-- --</div>
  </div>
</div>
<!-- Clinical Snapshot -->
<div class="snap-hdr">
  <div>
    <h2 id="snap-date">Loading...</h2>
  </div>
  <div style="display:flex;align-items:center;gap:12px">
    <div style="display:flex;background:rgba(255,255,255,.12);border:1px solid rgba(255,255,255,.2);border-radius:8px;overflow:hidden">
      <button id="btn-today" onclick="switchSnapView('today')"
        style="padding:6px 16px;font-size:13px;font-weight:600;border:none;cursor:pointer;background:#fde68a;color:#78350f;transition:all .2s;font-family:inherit">
        Today
      </button>
      <button id="btn-tomorrow" onclick="switchSnapView('tomorrow')"
        style="padding:6px 16px;font-size:13px;font-weight:600;border:none;cursor:pointer;background:transparent;color:rgba(255,255,255,.7);transition:all .2s;font-family:inherit">
        Tomorrow
      </button>
    </div>
    <div style="font-size:12px;color:rgba(255,255,255,.5);font-weight:500">Auto-refreshes every 5 min</div>
  </div>
</div>

<!-- Stats Bar -->
<div class="stats-bar" id="stats-bar">
  <div class="stat-box"><div class="stat-num">--</div><div class="stat-lbl">Patients Today</div></div>
</div>


<!-- Day Summary Panel (hidden until day_mode === summary) -->
<div id="day-summary-section" class="day-summary-wrap" style="display:none">
  <div class="day-summary-banner">
    <div>
      <h3>&#10003; Day Complete</h3>
      <p id="day-summary-date"></p>
    </div>
    <div style="text-align:right">
      <div style="font-size:28px;font-weight:800;color:#34d399" id="day-prod-num">--</div>
      <div style="font-size:10px;color:#4b7a5e;text-transform:uppercase;letter-spacing:.6px">Day Production</div>
    </div>
  </div>
  <div class="day-stats-row" id="day-stats-row"></div>
  <div class="day-prov-grid" id="day-prov-grid"></div>
  <div class="cancelled-list" id="cancelled-list-wrap" style="display:none">
    <div class="cancelled-hdr">Cancellations &amp; No-Shows</div>
    <div id="cancelled-list-body"></div>
  </div>
  <div class="cancelled-list" id="rescheduled-list-wrap" style="display:none;margin-top:10px">
    <div class="cancelled-hdr" style="color:#60a5fa">&#8635; Rescheduled Mid-Day</div>
    <div id="rescheduled-list-body"></div>
  </div>
  <div class="cancelled-list" id="deleted-list-wrap" style="display:none;margin-top:10px">
    <div class="cancelled-hdr" style="color:#f59e0b">&#9888; Deleted From Schedule</div>
    <div id="deleted-list-body"></div>
  </div>
  <div id="missing-slips-wrap" style="display:none;margin-top:10px;background:#ffffff;border:1px solid #fed7aa;border-radius:10px;overflow:hidden">
    <div id="missing-slips-hdr" onclick="toggleMissingSlips()"
      style="padding:12px 16px;cursor:pointer;display:flex;justify-content:space-between;align-items:center;user-select:none">
      <div style="font-size:12px;font-weight:700;color:#f59e0b">&#9888; Missing Routing Slips &nbsp;<span id="missing-slips-count" style="background:rgba(245,158,11,.2);padding:1px 8px;border-radius:20px;font-size:11px"></span></div>
      <div id="missing-slips-chevron" style="color:#f59e0b;font-size:14px;transition:transform .2s">&#9660;</div>
    </div>
    <div id="missing-slips-body" style="display:none;border-top:1px solid #2d1f0e"></div>
  </div>
</div>

<!-- Provider Schedule Cards -->
<div class="prov-row" id="prov-row">
  <div style="color:#94a3b8;font-size:13px;padding:20px">Loading schedule...</div>
</div>

<!-- Alert Grid -->
<div class="alert-grid" id="alert-grid">
  <div style="color:#94a3b8;font-size:13px;padding:20px">Loading alerts...</div>
</div>

<!-- Financial Stats Section -->
<div id="financial-section" style="background:#f0f2f7;border-bottom:1px solid #e2e8f0;padding:18px 28px">

  <!-- Date range control -->
  <div style="display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:10px;margin-bottom:16px">
    <div style="font-size:11px;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:.8px">Financial Overview</div>
    <div style="display:flex;align-items:center;gap:8px;flex-wrap:wrap">
      <div style="display:flex;gap:4px" id="fin-presets">
        <button onclick="finPreset('mtd',this)" class="fin-preset active-preset" style="padding:5px 12px;font-size:11px;border:1px solid #bfdbfe;border-radius:6px;background:#dbeafe;color:#1d4ed8;cursor:pointer">MTD</button>
        <button onclick="finPreset('l30',this)" class="fin-preset" style="padding:5px 12px;font-size:11px;border:1px solid #e2e8f0;border-radius:6px;background:#ffffff;color:#64748b;cursor:pointer">Last 30d</button>
        <button onclick="finPreset('l90',this)" class="fin-preset" style="padding:5px 12px;font-size:11px;border:1px solid #e2e8f0;border-radius:6px;background:#ffffff;color:#64748b;cursor:pointer">Last 90d</button>
        <button onclick="finPreset('ytd',this)" class="fin-preset" style="padding:5px 12px;font-size:11px;border:1px solid #e2e8f0;border-radius:6px;background:#ffffff;color:#64748b;cursor:pointer">YTD</button>
      </div>
      <div style="display:flex;align-items:center;gap:6px;font-size:12px;color:#64748b">
        <input type="date" id="fin-from" style="font-size:11px;padding:4px 8px;border-radius:6px;border:1px solid #e2e8f0;background:#ffffff;color:#1e293b;width:120px">
        <span>to</span>
        <input type="date" id="fin-to" style="font-size:11px;padding:4px 8px;border-radius:6px;border:1px solid #e2e8f0;background:#ffffff;color:#1e293b;width:120px">
        <button onclick="loadFinancial()" style="padding:5px 12px;font-size:11px;border:1px solid #bfdbfe;border-radius:6px;background:#dbeafe;color:#1d4ed8;cursor:pointer">Load</button>
      </div>
    </div>
  </div>

  <div id="fin-loading" style="color:#64748b;font-size:13px;padding:20px 0">Loading financial data...</div>
  <div id="fin-content" style="display:none">

    <!-- Production vs Collections -->
    <div style="background:#ffffff;border:1px solid #e2e8f0;border-radius:10px;padding:16px;margin-bottom:14px">
      <div style="font-size:11px;color:#64748b;text-transform:uppercase;letter-spacing:.8px;margin-bottom:12px">Production vs collections</div>
      <div style="display:grid;grid-template-columns:repeat(5,1fr);gap:10px;margin-bottom:16px">
        <div style="background:#f8fafc;border-radius:8px;padding:12px;border:1px solid #e2e8f0">
          <div style="font-size:11px;color:#64748b;margin-bottom:4px">Production</div>
          <div style="font-size:20px;font-weight:700;color:#1e293b" id="f-prod">--</div>
          <div style="font-size:10px;color:#94a3b8;margin-top:3px" id="f-procs">-- procedures</div>
        </div>
        <div style="background:#f8fafc;border-radius:8px;padding:12px;border:1px solid #e2e8f0">
          <div style="font-size:11px;color:#64748b;margin-bottom:4px">Collected</div>
          <div style="font-size:20px;font-weight:700;color:#34d399" id="f-coll">--</div>
          <div style="font-size:10px;color:#94a3b8;margin-top:3px" id="f-coll-sub">pat + ins</div>
        </div>
        <div style="background:#f8fafc;border-radius:8px;padding:12px;border:1px solid #e2e8f0">
          <div style="font-size:11px;color:#64748b;margin-bottom:4px">PPO collection rate</div>
          <div style="font-size:20px;font-weight:700" id="f-ppo-rate">--</div>
          <div style="font-size:10px;color:#94a3b8;margin-top:3px">target 95%+</div>
        </div>
        <div style="background:#f8fafc;border-radius:8px;padding:12px;border:1px solid #e2e8f0">
          <div style="font-size:11px;color:#64748b;margin-bottom:4px">Medicaid rate</div>
          <div style="font-size:20px;font-weight:700;color:#94a3b8" id="f-med-rate">--</div>
          <div style="font-size:10px;color:#94a3b8;margin-top:3px">informational</div>
        </div>
        <div style="background:#f8fafc;border-radius:8px;padding:12px;border:1px solid #e2e8f0">
          <div style="font-size:11px;color:#64748b;margin-bottom:4px">Write-offs</div>
          <div style="font-size:20px;font-weight:700;color:#fbbf24" id="f-writeoffs">--</div>
          <div style="font-size:10px;color:#94a3b8;margin-top:3px">ins + adjustments</div>
        </div>
      </div>
      <!-- Provider summary -->
      <div style="font-size:10px;color:#64748b;text-transform:uppercase;letter-spacing:.8px;margin-bottom:8px">Avg collection per patient by provider</div>
      <div id="f-prov-list"></div>
    </div>

    <!-- AR Section -->
    <div style="background:#ffffff;border:1px solid #e2e8f0;border-radius:10px;padding:16px;margin-bottom:14px">
      <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px">
        <div style="font-size:11px;color:#64748b;text-transform:uppercase;letter-spacing:.8px">Outstanding insurance AR</div>
        <div style="font-size:10px;color:#94a3b8">Last 12 months only</div>
      </div>
      <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:14px">
        <div style="background:#f8fafc;border-radius:8px;padding:12px;border:1px solid #e2e8f0">
          <div style="font-size:11px;color:#64748b;margin-bottom:4px">0-30 days</div>
          <div style="font-size:18px;font-weight:700;color:#34d399" id="ar-0-30">--</div>
          <div style="font-size:10px;color:#94a3b8;margin-top:3px" id="ar-0-30-cnt">-- claims</div>
        </div>
        <div style="background:#f8fafc;border-radius:8px;padding:12px;border:1px solid #e2e8f0">
          <div style="font-size:11px;color:#64748b;margin-bottom:4px">31-60 days</div>
          <div style="font-size:18px;font-weight:700;color:#fbbf24" id="ar-31-60">--</div>
          <div style="font-size:10px;color:#94a3b8;margin-top:3px" id="ar-31-60-cnt">-- claims</div>
        </div>
        <div style="background:#f8fafc;border-radius:8px;padding:12px;border:1px solid #e2e8f0">
          <div style="font-size:11px;color:#64748b;margin-bottom:4px">61-90 days</div>
          <div style="font-size:18px;font-weight:700;color:#fb923c" id="ar-61-90">--</div>
          <div style="font-size:10px;color:#94a3b8;margin-top:3px" id="ar-61-90-cnt">-- claims</div>
        </div>
        <div style="background:#f8fafc;border-radius:8px;padding:12px;border:1px solid #e2e8f0">
          <div style="font-size:11px;color:#64748b;margin-bottom:4px">90+ days</div>
          <div style="font-size:18px;font-weight:700;color:#f87171" id="ar-90">--</div>
          <div style="font-size:10px;color:#94a3b8;margin-top:3px" id="ar-90-cnt">-- claims</div>
        </div>
      </div>
      <table style="width:100%;border-collapse:collapse;font-size:12px">
        <thead>
          <tr style="border-bottom:1px solid #e2e8f0">
            <th style="text-align:left;color:#64748b;font-weight:500;padding:6px 0;width:35%">Carrier</th>
            <th style="text-align:right;color:#64748b;font-weight:500;padding:6px 0;width:12%">Claims</th>
            <th style="text-align:right;color:#64748b;font-weight:500;padding:6px 0;width:18%">Outstanding</th>
            <th style="text-align:right;color:#64748b;font-weight:500;padding:6px 0;width:12%">Oldest</th>
            <th style="text-align:left;color:#64748b;font-weight:500;padding:6px 8px;width:23%">Age</th>
          </tr>
        </thead>
        <tbody id="ar-table"></tbody>
      </table>
    </div>

    <!-- New patients + carrier payments -->
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:14px">

      <!-- New patients pie -->
      <div style="background:#ffffff;border:1px solid #e2e8f0;border-radius:10px;padding:16px">
        <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px">
          <div style="font-size:11px;color:#64748b;text-transform:uppercase;letter-spacing:.8px">New patients</div>
          <div style="font-size:22px;font-weight:700;color:#1e293b" id="f-np-total">--</div>
        </div>
        <div style="display:flex;gap:12px;align-items:flex-start">
          <canvas id="np-pie" width="156" height="156" style="flex-shrink:0"></canvas>
          <div style="flex:1;min-width:0">
            <div id="np-legend" style="font-size:11px;color:#64748b;margin-bottom:10px"></div>
            <div id="np-ppo-list" style="margin-bottom:8px"></div>
            <div id="np-med-list"></div>
          </div>
        </div>
      </div>

      <!-- Patient payments by carrier -->
      <div style="background:#ffffff;border:1px solid #e2e8f0;border-radius:10px;padding:16px">
        <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px">
          <div style="font-size:11px;color:#64748b;text-transform:uppercase;letter-spacing:.8px">Patient payments by carrier</div>
          <div style="font-size:10px;color:#94a3b8">min. 3 patients</div>
        </div>
        <table style="width:100%;border-collapse:collapse;font-size:12px;table-layout:fixed">
          <thead>
            <tr style="border-bottom:1px solid #e2e8f0">
              <th style="text-align:left;color:#64748b;font-weight:500;padding:5px 0;width:40%">Carrier</th>
              <th style="text-align:right;color:#64748b;font-weight:500;padding:5px 0;width:15%">Pts</th>
              <th style="text-align:right;color:#64748b;font-weight:500;padding:5px 0;width:25%">Total</th>
              <th style="text-align:right;color:#64748b;font-weight:500;padding:5px 0;width:20%">Avg/pt</th>
            </tr>
          </thead>
          <tbody id="f-carrier-pay"></tbody>
        </table>
      </div>
    </div>

  </div>
</div>

<!-- Insurance Lookup Section -->
<div id="ins-lookup-section" style="background:#f0f2f7;border-bottom:1px solid #e2e8f0;padding:18px 28px">
  <div style="font-size:11px;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:.8px;margin-bottom:12px">
    Insurance Lookup &mdash; Search by Employer or Group Name
  </div>
  <div style="display:flex;gap:10px;align-items:center;max-width:520px;margin-bottom:14px">
    <input id="ins-search-input" type="text" placeholder="Type employer or group name (e.g. SEPTA, Walmart, Local 32...)"
      style="flex:1;padding:10px 14px;font-size:13px;border-radius:8px;border:1px solid #e2e8f0;
             background:#ffffff;color:#1e293b;outline:none"
      oninput="insLookupDebounce(this.value)"
      onkeydown="if(event.key==='Enter')insLookup(this.value)">
    <button onclick="insLookup(document.getElementById('ins-search-input').value)"
      style="padding:10px 18px;font-size:13px;font-weight:600;border:none;border-radius:8px;
             background:#2563eb;color:#fff;cursor:pointer">Search</button>
    <button onclick="document.getElementById('ins-search-input').value='';document.getElementById('ins-lookup-results').innerHTML=''"
      style="padding:10px 14px;font-size:13px;border:none;border-radius:8px;
             background:#f1f5f9;color:#64748b;cursor:pointer">Clear</button>
  </div>
  <div id="ins-lookup-results"></div>
</div>

<!-- Automation Scripts Section -->
<div class="scripts-section">
  <div class="scripts-lbl">Automation Scripts</div>
  <div class="ctrl" style="padding:0 0 12px 0;">
    <button class="btn bg" id="btnStart">Start All Scripts</button>
    <button class="btn br" id="btnStop">Stop All Scripts</button>
  </div>
</div>

<div class="wrap">
  <div class="scripts" id="scripts"></div>
  <div class="lpanel">
    <div class="lh">Activity Log</div>
    <div class="lb" id="log"><div class="lem">No activity yet today</div></div>
  </div>
</div>
<script>
var S = __SCRIPTS__;

function cd(nr, paused) {
  if (paused) return '<span style="color:#6e7681">Paused</span>';
  if (!nr || nr === 'Scheduled' || nr === 'Not scheduled') return nr || 'Scheduled';
  var p = nr.split(' '), now = new Date(), t = new Date();
  var tp = p[0].split(':'), h = parseInt(tp[0]), m = parseInt(tp[1]);
  if (p[1]==='PM' && h!==12) h+=12;
  if (p[1]==='AM' && h===12) h=0;
  t.setHours(h,m,0,0);
  var dn = p.slice(2).join(' ');
  if (dn) {
    var ds=['Sunday','Monday','Tuesday','Wednesday','Thursday','Friday','Saturday'];
    var td=ds.indexOf(dn), diff=(td-now.getDay()+7)%7;
    if (diff===0 && t<now) diff=7;
    t.setDate(t.getDate()+diff);
  } else if (t<now) t.setDate(t.getDate()+1);
  var mins=Math.round((t-now)/60000);
  if (mins<1) return '<b style="color:#2ea043">Running soon</b>';
  if (mins<60) return '<b style="color:#f0883e">in '+mins+'m</b>';
  return '<b style="color:#58a6ff">in '+Math.floor(mins/60)+'h '+(mins%60)+'m</b>';
}

function render(d) {
  var g=document.getElementById('scripts'); g.innerHTML='';
  Object.keys(S).forEach(function(sid){
    var sc=S[sid], st=(d.scripts&&d.scripts[sid])||{};
    var paused=st.paused||false, running=st.running||false;
    var badge=running?'<span class="badge brun">Running</span>':
              paused?'<span class="badge bp">Paused</span>':
                     '<span class="badge ba">Active</span>';
    var spin=running?'<span class="spin"></span>':'';
    var cls=running?'running':paused?'paused':'';
    var lr=st.last_run||'Never', ls=st.last_status||'-', lsum=st.last_summary||'Not yet run';
    var sc2=(ls==='Success')?'ok':(ls==='-')?'':'er';
    g.innerHTML+='<div class="card '+cls+'" id="card-'+sid+'">'
      +'<div class="ch"><div class="ct">'
      +'<div class="ci">'+sc.icon+'</div>'
      +'<div><div class="cn">'+sc.name+'</div><div class="cd">'+sc.desc+'</div></div>'
      +'</div><div class="cbtns">'+spin+badge
      +'<button class="btn bsm '+(paused?'bg':'bo')+'" onclick="tog(\''+sid+'\')" '+(running?'disabled':'')+'>'+( paused?'Resume':'Pause')+'</button>'
      +'<button class="btn bsm bb" onclick="tst(\''+sid+'\')" '+(running?'disabled':'')+'>Test</button>'
      +'</div></div>'
      +'<div class="cb">'
      +'<div class="info"><label>Schedule</label><p>'+sc.cron+'</p></div>'
      +'<div class="info"><label>Next Run</label><p>'+cd(st.next_run,paused)+'</p></div>'
      +'<div class="info"><label>Last Run</label><p>'+lr+'</p></div>'
      +'<div class="info"><label>Status</label><p class="'+sc2+'">'+ls+'</p></div>'
      +'<div class="sum"><span style="font-size:10px;color:#555">LAST SUMMARY</span><br>'
      +'<span class="'+sc2+'">'+lsum+'</span></div>'
      +'</div></div>';
  });
}

function render_log(l) {
  var el=document.getElementById('log');
  if (!l||!l.length){el.innerHTML='<div class="lem">No activity yet today</div>';return;}
  el.innerHTML=l.map(function(e){
    return '<div class="li"><div class="lt">'+e.time+'</div>'
      +'<div class="ls" style="color:'+e.color+'">'+e.icon+' '+e.script+' - '+e.status+'</div>'
      +'<div style="color:#8b949e;font-size:11px">'+e.summary+'</div></div>';
  }).join('');
}

function go() {
  fetch('/api/state').then(function(r){return r.json();}).then(function(d){render(d);render_log(d.activity_log);}).catch(function(e){console.error(e);});
}
function tog(sid) { fetch('/api/toggle/'+sid,{method:'POST'}).then(go); }
function tst(sid) { fetch('/api/test/'+sid,{method:'POST'}).then(function(){setTimeout(go,500);}); }

document.getElementById('btnStart').onclick = function(){ fetch('/api/start_all',{method:'POST'}).then(go); };
document.getElementById('btnStop').onclick  = function(){ fetch('/api/stop_all', {method:'POST'}).then(go); };

// ============================================================
// CLINICAL SNAPSHOT
// ============================================================

function risk_flag(risk, bad) {
  if (bad === 0) return '';
  var cls = risk === 'high' ? 'flag-red' : 'flag-yellow';
  return '<span class="flag ' + cls + '">' + bad + ' no-show' + (bad>1?'s':'') + '</span>';
}

function apt_flags(apt) {
  var f = '';
  if (apt.is_new) f += '<span class="flag flag-green">NEW</span>';
  if (apt.prov_changed && apt.prev_prov)
    f += '<span class="flag flag-yellow" title="Last seen by '+apt.prev_prov+'">PROV CHANGE</span>';
  if (apt.copay) f += '<span class="flag flag-blue">'+apt.copay+'</span>';
  f += risk_flag(apt.noshow_risk, apt.noshow_bad);
  return f;
}

function render_snapshot(d) {
  if (!d || d.error) return;

  // Date header
  var hdr = d.date;
  if (d.is_preview) hdr += ' <span class="preview-badge">MONDAY PREVIEW</span>';
  document.getElementById('snap-date').innerHTML = hdr;

  // Stats bar — 5 boxes only
  var al = d.alerts;
  var unconf = al.unconfirmed ? al.unconfirmed.length : 0;
  var insv   = al.ins_verify  ? al.ins_verify.length  : 0;
  var copay  = al.copay       ? al.copay.length        : 0;
  var prauth = al.prior_auth  ? al.prior_auth.length   : 0;

  document.getElementById('stats-bar').innerHTML =
    stat_box(d.total_pts, 'Patients Today', d.total_pts > 0 ? 'ok' : 'neutral') +
    stat_box(unconf, 'Unconfirmed',   unconf > 0 ? 'danger'  : 'ok') +
    stat_box(insv,   'Ins. Verify',   insv > 0   ? 'warn'    : 'ok') +
    stat_box(copay,  'Co-Pay',        copay > 0  ? 'warn'    : 'ok') +
    stat_box(prauth, 'Prior Auth',    prauth > 0 ? 'danger'  : 'ok');

  // Provider cards
  var cards = d.prov_cards;
  var prow  = document.getElementById('prov-row');
  if (!cards || Object.keys(cards).length === 0) {
    prow.innerHTML = '<div style="color:#94a3b8;padding:20px;font-size:13px">No providers scheduled today</div>';
  } else {
    var html = '';
    var sorted_provs = Object.keys(cards).sort(function(a,b){
      return cards[a].name.localeCompare(cards[b].name);
    });
    sorted_provs.forEach(function(pid) {
      var pc = cards[pid];
      var h  = pc.health;
      var apt_html = '';
      if (pc.apts && pc.apts.length > 0) {
        pc.apts.forEach(function(apt) {
          var codes_str = apt.codes.slice(0,3).join(', ') + (apt.codes.length > 3 ? '...' : '');
          var lv = apt.last_visit || '';
          var room = apt.room ? ' | ' + apt.room : '';
          var s = apt.apt_status || 1;
          var row_cls = s === 2 ? 'apt-complete' : (s === 5 ? 'apt-broken' : (s === 3 ? 'apt-broken' : ''));
          var dot_cls = s === 2 ? 'dot-complete' : (s === 5 ? 'dot-broken' : (s === 3 ? 'dot-unscheduled' : 'dot-scheduled'));
          var status_label = s === 2 ? ' <span class="flag flag-green" style="font-size:9px">DONE</span>'
                           : s === 5 ? ' <span class="flag flag-red" style="font-size:9px">NO-SHOW</span>'
                           : s === 3 ? ' <span class="flag flag-yellow" style="font-size:9px">CANCELLED</span>'
                           : '';
          var slip_badge = '';
          if (s === 2) {
            slip_badge = apt.routing_slip
              ? '<div style="margin-top:3px"><span style="display:inline-block;padding:1px 7px;border-radius:20px;font-size:9px;font-weight:700;background:#dcfce7;color:#15803d;border:1px solid #86efac">&#10003; Slip Scanned</span></div>'
              : '<div style="margin-top:3px"><span style="display:inline-block;padding:1px 7px;border-radius:20px;font-size:9px;font-weight:700;background:#fee2e2;color:#dc2626;border:1px solid #fca5a5">&#9888; Slip Missing</span></div>';
          }
          apt_html +=
            '<div class="apt-row ' + row_cls + '">' +
            '<div class="apt-status-dot ' + dot_cls + '"></div>' +
            '<div class="apt-time">' + apt.time + '</div>' +
            '<div class="apt-info">' +
              '<div class="apt-patient">' + apt.patient + status_label + '</div>' +
              '<div class="apt-meta">' + codes_str + room + (lv ? ' | ' + lv : '') + '</div>' +
              slip_badge +
              '<div class="apt-flags">' + (s < 3 ? apt_flags(apt) : '') + '</div>' +
            '</div></div>';
        });
      } else {
        apt_html = '<div class="alert-empty">No appointments</div>';
      }

      var health_html = '';
      if (h && h.score !== null) {
        var pct = h.score;
        var risk_cls = h.high_risk && h.high_risk.length > 0 ? 'bad' : 'good';
        var risk_txt = h.high_risk && h.high_risk.length > 0 ? h.high_risk.length+' high risk' : 'Low';
        var prod_cls = h.has_high_val ? 'good' : '';
        var prod_txt = h.has_high_val ? 'High Value' : 'Routine';
        health_html =
          '<div class="health-section">' +
          '<div class="health-top">' +
            '<span class="health-label">Schedule Health</span>' +
            '<span class="health-score" style="color:'+h.color+'">' + h.score + '</span>' +
          '</div>' +
          '<div class="health-bar"><div class="health-fill" style="width:'+pct+'%;background:'+h.color+'"></div></div>' +
          '<div class="health-rating" style="color:'+h.color+'">' + h.rating + '</div>' +
          '<div class="health-grid">' +
            '<div class="hd-item"><span>Utilization</span><span class="hd-val">'+h.util_rate+'%</span></div>' +
            '<div class="hd-item"><span>Confirmed</span><span class="hd-val">'+h.conf_rate+'%</span></div>' +
            '<div class="hd-item"><span>No-Show</span><span class="hd-val '+risk_cls+'">'+risk_txt+'</span></div>' +
            '<div class="hd-item"><span>Production</span><span class="hd-val '+prod_cls+'">'+prod_txt+'</span></div>' +
            '<div class="hd-item"><span>New Patients</span><span class="hd-val">'+h.new_pts+'</span></div>' +
            '<div class="hd-item"><span>PPO+Self Pay</span><span class="hd-val '+(h.payer_pct >= 50 ? 'good' : h.payer_pct >= 35 ? '' : 'bad')+'">'+(h.payer_pct||0)+'%</span></div>' +
          '</div>' +
          (h.open_slots.length > 0 ?
            '<div class="open-slots"><div class="open-lbl">Open Chair Time</div>' +
            h.open_slots.map(function(s){ return '<div class="open-slot">&#9670; '+s+'</div>'; }).join('') +
            '</div>' : '') +
          '</div>';
      }

      html +=
        '<div class="prov-card">' +
        '<div class="prov-card-hdr">' +
          '<div><div class="prov-name">' + pc.name + '</div>' +
          '<div class="prov-meta">' + pc.apts.length + ' patients' +
          (pc.hours && pc.hours.trim() !== '–' ? ' &nbsp;&middot;&nbsp; ' + pc.hours : '') +
          '</div></div>' +
        '</div>' +
        '<div class="apt-list">' + apt_html + '</div>' +
        health_html +
        '</div>';
    });
    prow.innerHTML = html;
  }

  // Alert grid
  render_alerts(d.alerts);
}

function stat_box(n, lbl, cls) {
  return '<div class="stat-box ' + (cls||'') + '"><div class="stat-num">' + n +
         '</div><div class="stat-lbl">' + lbl + '</div></div>';
}

function alert_col(title, color, border_color, items, fields) {
  var cnt = items && items.length > 0 ? ' <span style="opacity:.7;font-weight:400">('+items.length+')</span>' : '';
  var body = '';
  if (!items || items.length === 0) {
    body = '<div class="alert-empty">All clear</div>';
  } else {
    items.forEach(function(it) {
      body += '<div class="alert-item" style="border-left-color:'+border_color+'">' +
        '<div class="alert-time">' + it.time + '</div>' +
        '<div class="alert-pt">' + it.patient + '</div>' +
        '<div class="alert-sub">' + fields(it) + '</div>' +
        '</div>';
    });
  }
  return '<div class="alert-col">' +
    '<div class="alert-col-hdr" style="color:'+color+';border-left-color:'+border_color+'">' +
      title + cnt +
    '</div>' +
    '<div class="alert-col-body">' + body + '</div></div>';
}

function render_alerts(al) {
  var grid = document.getElementById('alert-grid');
  var html = '';

  html += alert_col('Unconfirmed', '#fb923c', '#fb923c',
    al.unconfirmed, function(it) {
      return '<span style="color:#94a3b8">' + it.provider + '</span>' +
             '<br>' + it.phone +
             '<br><span style="color:#fb923c;font-style:italic">' + it.status + '</span>';
    });

  html += alert_col('Ins. Verify', '#a78bfa', '#a78bfa',
    al.ins_verify, function(it) {
      var c = it.color === 'red' ? '#f87171' : '#fbbf24';
      return '<span style="color:#94a3b8">' + it.provider + '</span>' +
             '<br>' + it.carrier +
             '<br><span style="color:'+c+';font-weight:600">' + it.label + '</span>';
    });

  html += alert_col('Co-Pay', '#60a5fa', '#60a5fa',
    al.copay, function(it) {
      var amt_color = it.amount === 'Check copay' ? '#94a3b8' : '#34d399';
      return '<span style="color:#94a3b8">' + it.provider + '</span>' +
             '<br><span style="color:'+amt_color+';font-weight:600">' + it.amount + '</span>';
    });

  html += alert_col('Prior Auth', '#f87171', '#f87171',
    al.prior_auth, function(it) {
      return '<span style="color:#94a3b8">' + it.provider + '</span>' +
             '<br><span style="color:#f87171;font-weight:600">' + it.codes + '</span>' +
             '<br><span style="color:#64748b">' + it.carrier + '</span>';
    });

  grid.innerHTML = html;
}


function fmt_money_k(n) {
  if (n >= 1000) return '$' + (n/1000).toFixed(1) + 'k';
  return '$' + n.toLocaleString();
}

function render_day_summary(ds, date_str) {
  if (!ds) return;
  var section = document.getElementById('day-summary-section');
  var prow    = document.getElementById('prov-row');

  if (ds.mode === 'summary') {
    section.style.display = 'block';
    prow.style.display    = 'none';
    document.getElementById('day-summary-date').textContent = date_str || '';
    document.getElementById('day-prod-num').textContent = fmt_money_k(ds.day_prod || 0);

    // Stats row
    var sr = document.getElementById('day-stats-row');
    sr.innerHTML =
      day_stat(ds.completed,    'Patients Seen',     '#34d399') +
      day_stat(ds.cancelled,    'Cancelled/No-Show', '#f87171') +
      (ds.rescheduled > 0 ? day_stat(ds.rescheduled, 'Rescheduled', '#60a5fa') : '') +
      (ds.deleted > 0     ? day_stat(ds.deleted,     'Deleted',     '#f59e0b') : '') +
      day_stat(ds.day_procs,    'Procedures Done',   '#94a3b8');

    // Provider grid
    var pg = document.getElementById('day-prov-grid');
    pg.innerHTML = (ds.prov_day || []).map(function(p) {
      return '<div class="day-prov-card">' +
        '<div class="day-prov-name">' + p.name + '</div>' +
        '<div class="day-prov-row"><span>Seen</span><span class="day-prov-val" style="color:#34d399">' + p.seen + '</span></div>' +
        (p.scheduled > 0 ? '<div class="day-prov-row"><span>Still Scheduled</span><span class="day-prov-val" style="color:#60a5fa">' + p.scheduled + '</span></div>' : '') +
        (p.cancelled > 0 ? '<div class="day-prov-row"><span>Cancelled/No-show</span><span class="day-prov-val" style="color:#f87171">' + p.cancelled + '</span></div>' : '') +
        '</div>';
    }).join('');

    // Cancelled list
    var cl = ds.cancelled_list || [];
    var cw = document.getElementById('cancelled-list-wrap');
    var cb = document.getElementById('cancelled-list-body');
    if (cl.length > 0) {
      cw.style.display = 'block';
      cb.innerHTML = cl.map(function(c) {
        var badge = c.apt_status === 5
          ? '<span class="noshow-badge">No-Show</span>'
          : '<span class="unsched-badge">Cancelled</span>';
        var resched = c.rescheduled
          ? '<span class="resched-badge">&#10003; Rescheduled</span>'
          : '';
        return '<div class="cancelled-item">' +
          '<div>' +
            '<div class="cancelled-name">' + c.patient + '</div>' +
            '<div class="cancelled-meta">' + c.time + ' &nbsp;&middot;&nbsp; ' + c.provider +
            (c.phone ? ' &nbsp;&middot;&nbsp; ' + c.phone : '') + '</div>' +
          '</div>' +
          '<div style="display:flex;gap:6px;flex-shrink:0">' + badge + resched + '</div>' +
          '</div>';
      }).join('');
    } else {
      cw.style.display = 'none';
    }

    // Rescheduled mid-day list
    var rl = ds.rescheduled_list || [];
    var rw = document.getElementById('rescheduled-list-wrap');
    var rb = document.getElementById('rescheduled-list-body');
    if (rl.length > 0) {
      rw.style.display = 'block';
      rb.innerHTML = rl.map(function(r) {
        return '<div class="cancelled-item">' +
          '<div>' +
            '<div class="cancelled-name">' + r.patient + '</div>' +
            '<div class="cancelled-meta">' + r.time + ' &nbsp;&middot;&nbsp; ' + r.provider +
            (r.phone ? ' &nbsp;&middot;&nbsp; ' + r.phone : '') + '</div>' +
          '</div>' +
          '<div style="display:flex;gap:6px;flex-shrink:0">' +
            '<span class="resched-badge">&#8635; ' + (r.new_date || 'Future') + '</span>' +
          '</div>' +
        '</div>';
      }).join('');
    } else {
      rw.style.display = 'none';
    }

    // Deleted appointments list
    var dl = ds.deleted_list || [];
    var dw = document.getElementById('deleted-list-wrap');
    var db = document.getElementById('deleted-list-body');
    if (dl.length > 0) {
      dw.style.display = 'block';
      db.innerHTML = dl.map(function(d) {
        return '<div class="cancelled-item">' +
          '<div>' +
            '<div class="cancelled-name">' + d.patient + '</div>' +
            '<div class="cancelled-meta">' + d.time + ' &nbsp;&middot;&nbsp; ' + d.provider +
            (d.phone ? ' &nbsp;&middot;&nbsp; ' + d.phone : '') + '</div>' +
          '</div>' +
          '<div style="display:flex;gap:6px;flex-shrink:0">' +
            '<span style="padding:2px 8px;border-radius:20px;font-size:10px;font-weight:700;background:rgba(245,158,11,.15);color:#f59e0b;border:1px solid rgba(245,158,11,.3)">Deleted</span>' +
          '</div>' +
        '</div>';
      }).join('');
    } else {
      dw.style.display = 'none';
    }

    // Missing routing slips card
    var msl = ds.missing_slips_list || [];
    var msw = document.getElementById('missing-slips-wrap');
    var msb = document.getElementById('missing-slips-body');
    var msc = document.getElementById('missing-slips-count');
    if (msl.length > 0) {
      msw.style.display = 'block';
      msc.textContent   = msl.length + ' missing';
      msb.innerHTML = msl.map(function(m) {
        return '<div class="cancelled-item">' +
          '<div>' +
            '<div class="cancelled-name">' + m.patient + '</div>' +
            '<div class="cancelled-meta">' + m.time + ' &nbsp;&middot;&nbsp; ' + m.provider + '</div>' +
          '</div>' +
          '<span style="padding:1px 7px;border-radius:20px;font-size:9px;font-weight:700;background:rgba(248,113,113,.12);color:#f87171;border:1px solid rgba(248,113,113,.25)">&#9888; Slip Missing</span>' +
        '</div>';
      }).join('');
    } else {
      msw.style.display = 'none';
    }

  } else {
    section.style.display = 'none';
    prow.style.display    = 'flex';
  }
}

function toggleMissingSlips() {
  var body    = document.getElementById('missing-slips-body');
  var chevron = document.getElementById('missing-slips-chevron');
  var open    = body.style.display !== 'none';
  body.style.display    = open ? 'none' : 'block';
  chevron.style.transform = open ? '' : 'rotate(180deg)';
}

function day_stat(n, lbl, color) {
  return '<div class="day-stat"><div class="day-stat-num" style="color:' + color + '">' + n +
    '</div><div class="day-stat-lbl">' + lbl + '</div></div>';
}

var current_snap_view = 'today';

function switchSnapView(view) {
  current_snap_view = view;
  var btnToday    = document.getElementById('btn-today');
  var btnTomorrow = document.getElementById('btn-tomorrow');
  if (view === 'tomorrow') {
    btnToday.style.background    = 'transparent';
    btnToday.style.color         = 'rgba(255,255,255,.7)';
    btnTomorrow.style.background = '#fde68a';
    btnTomorrow.style.color      = '#92400e';
  } else {
    btnToday.style.background    = '#fde68a';
    btnToday.style.color         = '#92400e';
    btnTomorrow.style.background = 'transparent';
    btnTomorrow.style.color      = 'rgba(255,255,255,.7)';
  }
  fetch_snapshot();
}

function fetch_snapshot() {
  var url = current_snap_view === 'tomorrow' ? '/api/snapshot?view=tomorrow' : '/api/snapshot';
  fetch(url)
    .then(function(r) { return r.json(); })
    .then(function(d) { render_snapshot(d); render_day_summary(d.day_summary, d.date); })
    .catch(function(e) { console.error('Snapshot error:', e); });
}

setInterval(fetch_snapshot, 300000);
fetch_snapshot();

// ============================================================
// FINANCIAL STATS
// ============================================================

function fmt_money(n) {
  return '$' + Math.round(n).toLocaleString();
}

function fin_age_badge(days) {
  if (days <= 30)  return '<span style="padding:2px 8px;border-radius:20px;font-size:10px;font-weight:600;background:rgba(52,211,153,.15);color:#34d399;border:1px solid rgba(52,211,153,.2)">0-30d</span>';
  if (days <= 60)  return '<span style="padding:2px 8px;border-radius:20px;font-size:10px;font-weight:600;background:rgba(251,191,36,.15);color:#fbbf24;border:1px solid rgba(251,191,36,.2)">31-60d</span>';
  if (days <= 90)  return '<span style="padding:2px 8px;border-radius:20px;font-size:10px;font-weight:600;background:rgba(251,146,60,.15);color:#fb923c;border:1px solid rgba(251,146,60,.2)">61-90d</span>';
  return '<span style="padding:2px 8px;border-radius:20px;font-size:10px;font-weight:600;background:rgba(248,113,113,.15);color:#f87171;border:1px solid rgba(248,113,113,.2)">90+ days</span>';
}

function render_financial(d) {
  if (!d || d.error) {
    document.getElementById('fin-loading').textContent = 'Error loading financial data';
    return;
  }
  document.getElementById('fin-loading').style.display = 'none';
  document.getElementById('fin-content').style.display = 'block';

  // Production vs collections
  document.getElementById('f-prod').textContent      = fmt_money(d.total_production);
  document.getElementById('f-procs').textContent     = d.total_procedures + ' procedures';
  document.getElementById('f-coll').textContent      = fmt_money(d.total_collected);
  document.getElementById('f-coll-sub').textContent  = fmt_money(d.patient_collected) + ' pat / ' + fmt_money(d.ins_collected) + ' ins';
  document.getElementById('f-writeoffs').textContent = fmt_money(d.total_writeoffs);

  var ppoRate = d.ppo_coll_rate || 0;
  var ppoEl   = document.getElementById('f-ppo-rate');
  ppoEl.textContent = ppoRate + '%';
  ppoEl.style.color = ppoRate >= 95 ? '#34d399' : ppoRate >= 80 ? '#fbbf24' : '#f87171';

  var medRate = d.med_coll_rate || 0;
  var medEl = document.getElementById('f-med-rate');
  medEl.textContent = medRate + '%';

  // Provider summary
  var pl = document.getElementById('f-prov-list');
  pl.innerHTML = (d.prov_summary || []).map(function(p) {
    return '<div style="display:flex;justify-content:space-between;align-items:center;padding:7px 0;border-bottom:1px solid #f1f5f9">' +
      '<div style="font-size:13px;color:#1e293b;font-weight:500">' + p.name + '</div>' +
      '<div style="display:flex;gap:20px">' +
        '<div style="text-align:right"><div style="font-size:13px;font-weight:500;color:#1e293b">' + fmt_money(p.production) + '</div><div style="font-size:10px;color:#94a3b8">production</div></div>' +
        '<div style="text-align:right"><div style="font-size:13px;color:#334155;font-weight:500">' + p.patients + '</div><div style="font-size:10px;color:#94a3b8">patients</div></div>' +
        '<div style="text-align:right"><div style="font-size:13px;font-weight:700;color:#16a34a">' + fmt_money(p.avg_collection) + '</div><div style="font-size:10px;color:#94a3b8">avg collected/pt</div></div>' +
      '</div></div>';
  }).join('');

  // AR buckets
  var bk = d.ar_buckets || {};
  var ct = d.ar_counts  || {};
  document.getElementById('ar-0-30').textContent     = fmt_money(bk['0-30']  || 0);
  document.getElementById('ar-31-60').textContent    = fmt_money(bk['31-60'] || 0);
  document.getElementById('ar-61-90').textContent    = fmt_money(bk['61-90'] || 0);
  document.getElementById('ar-90').textContent       = fmt_money(bk['90+']   || 0);
  document.getElementById('ar-0-30-cnt').textContent  = (ct['0-30']  || 0) + ' claims';
  document.getElementById('ar-31-60-cnt').textContent = (ct['31-60'] || 0) + ' claims';
  document.getElementById('ar-61-90-cnt').textContent = (ct['61-90'] || 0) + ' claims';
  document.getElementById('ar-90-cnt').textContent    = (ct['90+']   || 0) + ' claims';

  // AR table with expandable rows
  var tb = document.getElementById('ar-table');
  tb.innerHTML = (d.ar_carriers || []).map(function(c, i) {
    var pts_html = c.patients.slice(0,10).map(function(p) {
      return '<tr style="background:#f8fafc">' +
        '<td style="padding:5px 0 5px 20px;color:#64748b;font-size:11px">' + p.name + '</td>' +
        '<td colspan="2"></td>' +
        '<td style="text-align:right;padding:5px 0;color:#64748b;font-size:11px">' + p.days + 'd</td>' +
        '<td style="padding:5px 8px;color:#64748b;font-size:11px">' + fmt_money(p.amount) + '</td>' +
      '</tr>';
    }).join('');
    return '<tr style="border-bottom:1px solid #1e2d3d;cursor:pointer" onclick="toggleAR(' + i + ')">' +
      '<td style="padding:8px 0;color:#1e293b;font-size:12px">' + c.carrier + '</td>' +
      '<td style="text-align:right;padding:8px 0;color:#64748b;font-size:12px">' + c.claims + '</td>' +
      '<td style="text-align:right;padding:8px 0;color:#1e293b;font-weight:600;font-size:12px">' + fmt_money(c.total) + '</td>' +
      '<td style="text-align:right;padding:8px 0;color:#64748b;font-size:12px">' + c.oldest + 'd</td>' +
      '<td style="padding:8px 8px;font-size:12px">' + fin_age_badge(c.oldest) + '</td>' +
    '</tr>' +
    '<tr id="ar-expand-' + i + '" style="display:none"><td colspan="5" style="padding:0"><table style="width:100%;border-collapse:collapse">' + pts_html + '</table></td></tr>';
  }).join('');

  // New patients pie
  draw_np_pie(d);

  // Carrier payments
  var cp = document.getElementById('f-carrier-pay');
  cp.innerHTML = (d.pay_by_carrier || []).map(function(r) {
    return '<tr style="border-bottom:1px solid #1e2d3d">' +
      '<td style="padding:7px 0;color:#1e293b;font-size:12px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">' + r.carrier + '</td>' +
      '<td style="text-align:right;padding:7px 0;color:#64748b;font-size:12px">' + r.patients + '</td>' +
      '<td style="text-align:right;padding:7px 0;color:#1e293b;font-weight:600;font-size:12px">' + fmt_money(r.total) + '</td>' +
      '<td style="text-align:right;padding:7px 0;color:#64748b;font-size:12px">' + fmt_money(r.avg) + '</td>' +
    '</tr>';
  }).join('');
}

function toggleAR(idx) {
  var row = document.getElementById('ar-expand-' + idx);
  if (row) row.style.display = row.style.display === 'none' ? 'table-row' : 'none';
}

function draw_np_pie(d) {
  var canvas = document.getElementById('np-pie');
  if (!canvas || !canvas.getContext) return;
  var ctx    = canvas.getContext('2d');
  var total  = d.new_patients || 1;
  document.getElementById('f-np-total').textContent = total;

  var slices = [
    {label:'PPO',        n: d.np_ppo,   color:'#378ADD'},
    {label:'Medicaid',   n: d.np_med,   color:'#1D9E75'},
    {label:'No ins',     n: d.np_none,  color:'#BA7517'},
    {label:'Other',      n: d.np_other, color:'#888780'},
  ].filter(function(s){ return s.n > 0; });

  var cx = 78, cy = 78, r = 65, hole = 36;
  var start = -Math.PI / 2;
  ctx.clearRect(0, 0, 156, 156);

  slices.forEach(function(s) {
    var angle = (s.n / total) * 2 * Math.PI;
    ctx.beginPath();
    ctx.moveTo(cx, cy);
    ctx.arc(cx, cy, r, start, start + angle);
    ctx.closePath();
    ctx.fillStyle = s.color;
    ctx.globalAlpha = 0.88;
    ctx.fill();
    start += angle;
  });

  ctx.globalAlpha = 1;
  ctx.beginPath();
  ctx.arc(cx, cy, hole, 0, 2 * Math.PI);
  ctx.fillStyle = '#111927';
  ctx.fill();

  // Center text
  ctx.fillStyle = '#e2e8f0';
  ctx.font = '600 14px sans-serif';
  ctx.textAlign = 'center';
  ctx.fillText(total, cx, cy + 5);

  // Legend — group totals
  var leg = document.getElementById('np-legend');
  leg.innerHTML = slices.map(function(s) {
    var pct = Math.round(s.n / total * 100);
    return '<div style="display:flex;align-items:center;gap:5px;margin-bottom:4px">' +
      '<span style="width:8px;height:8px;border-radius:2px;flex-shrink:0;background:' + s.color + ';opacity:.88;display:inline-block"></span>' +
      '<span style="color:#94a3b8">' + s.label + '</span>' +
      '<span style="color:#1e293b;font-weight:600;margin-left:auto">' + s.n + '</span>' +
      '<span style="color:#94a3b8;width:32px;text-align:right">' + pct + '%</span>' +
    '</div>';
  }).join('');

  // PPO carriers list
  var ppo_list = document.getElementById('np-ppo-list');
  if (d.np_ppo_carriers && d.np_ppo_carriers.length > 0) {
    ppo_list.innerHTML =
      '<div style="font-size:10px;color:#378ADD;text-transform:uppercase;letter-spacing:.5px;margin-bottom:4px">PPO breakdown</div>' +
      d.np_ppo_carriers.map(function(c) {
        var pct = Math.round(c.count / total * 100);
        return '<div style="display:flex;justify-content:space-between;font-size:11px;padding:2px 0;border-bottom:1px solid #1e2d3d">' +
          '<span style="color:#94a3b8;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;max-width:120px" title="'+c.name+'">' + c.name + '</span>' +
          '<span style="color:#378ADD;font-weight:500;flex-shrink:0;margin-left:4px">' + c.count + ' <span style="color:#94a3b8">(' + pct + '%)</span></span>' +
        '</div>';
      }).join('');
  } else {
    ppo_list.innerHTML = '';
  }

  // Medicaid carriers list
  var med_list = document.getElementById('np-med-list');
  if (d.np_med_carriers && d.np_med_carriers.length > 0) {
    med_list.innerHTML =
      '<div style="font-size:10px;color:#1D9E75;text-transform:uppercase;letter-spacing:.5px;margin-bottom:4px;margin-top:8px">Medicaid breakdown</div>' +
      d.np_med_carriers.map(function(c) {
        var pct = Math.round(c.count / total * 100);
        return '<div style="display:flex;justify-content:space-between;font-size:11px;padding:2px 0;border-bottom:1px solid #1e2d3d">' +
          '<span style="color:#94a3b8;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;max-width:120px" title="'+c.name+'">' + c.name + '</span>' +
          '<span style="color:#1D9E75;font-weight:500;flex-shrink:0;margin-left:4px">' + c.count + ' <span style="color:#94a3b8">(' + pct + '%)</span></span>' +
        '</div>';
      }).join('');
  } else {
    med_list.innerHTML = '';
  }
}

var _fin_today = new Date();
var _fin_m     = String(_fin_today.getMonth()+1).padStart(2,'0');
var _fin_d     = String(_fin_today.getDate()).padStart(2,'0');
var _fin_y     = _fin_today.getFullYear();
var _fin_today_str = _fin_y + '-' + _fin_m + '-' + _fin_d;
var _fin_month_str = _fin_y + '-' + _fin_m + '-01';

document.getElementById('fin-from').value = _fin_month_str;
document.getElementById('fin-to').value   = _fin_today_str;

function finPreset(key, btn) {
  document.querySelectorAll('.fin-preset').forEach(function(b){
    b.style.background = 'transparent';
    b.style.color = '#64748b';
    b.style.borderColor = '#1e2d3d';
  });
  btn.style.background   = '#1a2d45';
  btn.style.color        = '#60a5fa';
  btn.style.borderColor  = '#2d4a6b';

  var now  = new Date();
  var y    = now.getFullYear();
  var m    = now.getMonth();
  var d    = now.getDate();
  var from, to = _fin_today_str;

  if (key === 'mtd') {
    from = y + '-' + String(m+1).padStart(2,'0') + '-01';
  } else if (key === 'l30') {
    var f30 = new Date(now); f30.setDate(f30.getDate()-30);
    from = f30.getFullYear()+'-'+String(f30.getMonth()+1).padStart(2,'0')+'-'+String(f30.getDate()).padStart(2,'0');
  } else if (key === 'l90') {
    var f90 = new Date(now); f90.setDate(f90.getDate()-90);
    from = f90.getFullYear()+'-'+String(f90.getMonth()+1).padStart(2,'0')+'-'+String(f90.getDate()).padStart(2,'0');
  } else if (key === 'ytd') {
    from = y + '-01-01';
  }
  document.getElementById('fin-from').value = from;
  document.getElementById('fin-to').value   = to;
  loadFinancial();
}

function loadFinancial() {
  var from = document.getElementById('fin-from').value;
  var to   = document.getElementById('fin-to').value;
  document.getElementById('fin-loading').style.display = 'block';
  document.getElementById('fin-loading').textContent   = 'Loading...';
  document.getElementById('fin-content').style.display = 'none';
  fetch('/api/financial?from=' + from + '&to=' + to)
    .then(function(r){ return r.json(); })
    .then(function(d){ render_financial(d); })
    .catch(function(e){ document.getElementById('fin-loading').textContent = 'Error: ' + e; });
}

loadFinancial();


// ============================================================
// INSURANCE LOOKUP
// ============================================================

var _ins_debounce_timer = null;

function insLookupDebounce(val) {
  clearTimeout(_ins_debounce_timer);
  if (val.length < 2) {
    document.getElementById('ins-lookup-results').innerHTML = '';
    return;
  }
  _ins_debounce_timer = setTimeout(function(){ insLookup(val); }, 400);
}

function insLookup(q) {
  q = (q || '').trim();
  if (q.length < 2) return;
  var el = document.getElementById('ins-lookup-results');
  el.innerHTML = '<div style="color:#64748b;font-size:13px;padding:8px 0">Searching...</div>';
  fetch('/api/ins_lookup?q=' + encodeURIComponent(q))
    .then(function(r){ return r.json(); })
    .then(function(d){
      if (d.error) {
        el.innerHTML = '<div style="color:#f87171;font-size:12px;padding:8px 0">Error: ' + d.error + '</div>';
        return;
      }
      if (!d.results || d.results.length === 0) {
        el.innerHTML =
          '<div style="color:#64748b;font-size:13px;padding:12px 16px;background:#f8fafc;border-radius:8px;border:1px solid #e2e8f0">' +
          'No matches found for <b style="color:#1e293b">&quot;' + d.query + '&quot;</b></div>';
        return;
      }
      var html = '<div style="display:flex;flex-direction:column;gap:10px">';
      d.results.forEach(function(r) {
        var bc = r.match_type === 'Employer' ? '#1f6feb' : '#7c3aed';
        var pc = r.patient_count >= 5 ? '#34d399' : r.patient_count >= 2 ? '#fbbf24' : '#94a3b8';
        var samples = r.sample_patients.length > 0
          ? r.sample_patients.join(', ') + (r.patient_count > 5 ? ', ...' : '') : 'No samples';
        html +=
          '<div style="background:#ffffff;border:1px solid #e2e8f0;border-radius:10px;padding:14px 18px;display:flex;align-items:flex-start;gap:16px">' +
          '<div style="flex:1;min-width:0">' +
          '<div style="display:flex;align-items:center;gap:8px;margin-bottom:6px">' +
          '<span style="font-size:10px;font-weight:700;padding:2px 8px;border-radius:20px;background:' + bc + '22;color:' + bc + ';border:1px solid ' + bc + '44">' + r.match_type + '</span>' +
          '<span style="font-size:12px;color:#64748b">' + r.match_name + '</span></div>' +
          '<div style="font-size:18px;font-weight:700;color:#1e293b;margin-bottom:4px">' + r.carrier + '</div>' +
          (r.group_name ? '<div style="font-size:12px;color:#64748b">Group: <span style="color:#94a3b8">' + r.group_name + '</span>' + (r.group_num ? ' &nbsp;&middot;&nbsp; #' + r.group_num : '') + '</div>' : '') +
          '<div style="font-size:11px;color:#94a3b8;margin-top:6px">Seen with: ' + samples + '</div>' +
          '</div>' +
          '<div style="text-align:right;flex-shrink:0"><div style="font-size:28px;font-weight:800;color:' + pc + ';line-height:1">' + r.patient_count + '</div>' +
          '<div style="font-size:10px;color:#94a3b8;text-transform:uppercase;letter-spacing:.5px">patient' + (r.patient_count !== 1 ? 's' : '') + '</div></div>' +
          '</div>';
      });
      html += '</div>';
      el.innerHTML = html;
    })
    .catch(function(e){
      document.getElementById('ins-lookup-results').innerHTML =
        '<div style="color:#f87171;font-size:12px">Request failed: ' + e + '</div>';
    });
}

function tick(){
  var now=new Date(),h=now.getHours(),m=now.getMinutes(),ap=h>=12?'PM':'AM';
  h=h%12||12;
  document.getElementById('clk').textContent=h+':'+(m<10?'0':'')+m+' '+ap;
}
setInterval(go,10000); setInterval(tick,1000); tick(); go();
</script>
</body>
</html>"""

@app.route("/")
@login_required
def index():
    scripts_js = {}
    for sid, s in SCRIPTS.items():
        scripts_js[sid] = {"name":s["name"],"icon":s["icon"],"cron":s["cron"],"desc":s["desc"]}
    page = HTML.replace("__PRACTICE__", PRACTICE_NAME)
    page = page.replace("__SCRIPTS__", json.dumps(scripts_js))
    return page

@app.route("/api/state")
@login_required
def api_state():
    update_next_runs()
    return jsonify({"scripts": script_state, "activity_log": activity_log[:20]})

@app.route("/api/toggle/<sid>", methods=["POST"])
@login_required
def api_toggle(sid):
    if sid not in SCRIPTS: return jsonify({"error":"unknown"}), 404
    script_state[sid]["paused"] = not script_state[sid]["paused"]
    update_next_runs()
    return jsonify({"ok": True})

@app.route("/api/test/<sid>", methods=["POST"])
@login_required
def api_test(sid):
    if sid not in SCRIPTS: return jsonify({"error":"unknown"}), 404
    threading.Thread(target=run_script, args=(sid,), daemon=True).start()
    return jsonify({"ok": True})

@app.route("/api/start_all", methods=["POST"])
@login_required
def api_start_all():
    for sid in SCRIPTS: script_state[sid]["paused"] = False
    update_next_runs()
    return jsonify({"ok": True})

@app.route("/api/stop_all", methods=["POST"])
@login_required
def api_stop_all():
    for sid in SCRIPTS: script_state[sid]["paused"] = True
    update_next_runs()
    return jsonify({"ok": True})

# Track which dates have already been logged to avoid duplicate rows
_day_summary_logged = set()

def log_day_summary_to_excel(ds, snap_date_str):
    """Append one row to the daily summary Excel log. Safe to call multiple times — deduplicates by date."""
    global _day_summary_logged
    if snap_date_str in _day_summary_logged:
        return
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        log_path = os.path.join(BASE_DIR, "reports", "daily_summary_log.xlsx")
        os.makedirs(os.path.dirname(log_path), exist_ok=True)

        HEADERS = [
            "Date", "Day", "Patients Seen", "Cancelled/No-Show",
            "Rescheduled", "Deleted", "Day Production ($)", "Procedures Done",
            "Missing Routing Slips"
        ]

        # Load or create workbook
        if os.path.exists(log_path):
            wb = openpyxl.load_workbook(log_path)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Daily Summary"
            # Header row
            for col, h in enumerate(HEADERS, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font      = Font(bold=True, color="FFFFFF", size=11)
                cell.fill      = PatternFill("solid", fgColor="1E3A8A")
                cell.alignment = Alignment(horizontal="center")
            # Column widths
            widths = [14, 12, 16, 20, 14, 10, 20, 18, 22]
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        # Check if date already exists (in case of server restart)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and str(row[0]) == snap_date_str:
                _day_summary_logged.add(snap_date_str)
                return

        # Parse date for day-of-week
        from datetime import datetime as _dt
        dt = _dt.strptime(snap_date_str, "%Y-%m-%d")
        day_name = dt.strftime("%A")

        row_data = [
            snap_date_str,
            day_name,
            ds.get("completed", 0),
            ds.get("cancelled", 0),
            ds.get("rescheduled", 0),
            ds.get("deleted", 0),
            ds.get("day_prod", 0),
            ds.get("day_procs", 0),
            ds.get("missing_slips", 0),
        ]

        next_row = ws.max_row + 1
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=next_row, column=col, value=val)
            cell.alignment = Alignment(horizontal="center")
            # Alternate row shading
            if next_row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F0F4FF")

        # Bold the production column
        ws.cell(row=next_row, column=7).font = Font(bold=True)

        wb.save(log_path)
        _day_summary_logged.add(snap_date_str)
        log(f"Day summary logged to Excel for {snap_date_str}")
    except Exception as e:
        log(f"WARNING: Could not write day summary to Excel: {e}")



@app.route("/api/snapshot")
@login_required
def api_snapshot():
    try:
        import mysql.connector as mysql_connector
        from datetime import timedelta
        conn   = mysql_connector.connect(**DB_CONFIG)
        cursor = conn.cursor(dictionary=True)
        view   = request.args.get("view", "today")
        if view == "tomorrow":
            tomorrow = date.today() + timedelta(days=1)
            # If tomorrow is Sunday, show Monday instead
            if tomorrow.weekday() == 6:
                tomorrow = tomorrow + timedelta(days=1)
            data = get_snapshot_data(cursor, target_date=tomorrow)
        else:
            data = get_snapshot_data(cursor)
        cursor.close()
        conn.close()
        # Log end-of-day summary to Excel when day is complete
        if data.get("day_summary", {}).get("mode") == "summary":
            log_day_summary_to_excel(data["day_summary"], data.get("snap_date", ""))
        # Convert date objects to strings for JSON
        import json as _j
        return app.response_class(
            response=_j.dumps(data, default=str),
            mimetype="application/json"
        )
    except Exception as e:
        log(f"Snapshot API error: {e}")
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

# =============================================================================
# FINANCIAL STATS — Queries and API
# =============================================================================

MEDICAID_PLAN_CARRIERS = [
    "keystone first", "keystone mercy", "keystone first dentaquest",
    "keystone first vip", "health partner", "health partners",
    "health partner medicare", "aetna better health",
    "united healthcare community plan", "upmc", "avesis", "dentaquest",
]

def is_medicaid_carrier(carrier_name, plan_type):
    if plan_type == "f":
        return True
    if not carrier_name:
        return False
    cl = carrier_name.lower()
    return any(m in cl for m in MEDICAID_PLAN_CARRIERS)

def get_financial_stats(cursor, date_from, date_to):
    """Fetch all financial stats for the given date range."""

    # --- Production ---
    cursor.execute("""
        SELECT
            CONCAT(prov.FName,' ',prov.LName) AS ProviderName,
            prov.ProvNum,
            SUM(pl.ProcFee) AS Production,
            COUNT(DISTINCT pl.PatNum) AS Patients,
            COUNT(*) AS Procedures
        FROM procedurelog pl
        INNER JOIN provider prov ON prov.ProvNum = pl.ProvNum
        WHERE pl.ProcStatus = 2
        AND pl.ProcDate BETWEEN %s AND %s
        GROUP BY prov.ProvNum, ProviderName
        ORDER BY Production DESC
    """, (date_from, date_to))
    prod_rows = cursor.fetchall()

    total_production  = sum(float(r["Production"] or 0) for r in prod_rows)
    total_patients    = sum(int(r["Patients"] or 0) for r in prod_rows)
    total_procedures  = sum(int(r["Procedures"] or 0) for r in prod_rows)

    # --- Patient collections by provider (paysplit) ---
    cursor.execute("""
        SELECT
            ps.ProvNum,
            CONCAT(prov.FName,' ',prov.LName) AS ProviderName,
            SUM(ps.SplitAmt) AS PatientCollected
        FROM payment p
        INNER JOIN paysplit ps ON ps.PayNum = p.PayNum
        INNER JOIN provider prov ON prov.ProvNum = ps.ProvNum
        WHERE DATE(p.PayDate) BETWEEN %s AND %s
        AND p.PayAmt > 0
        AND ps.IsDiscount = 0
        AND p.PayType != 0
        AND ps.UnearnedType = 0
        GROUP BY ps.ProvNum, ProviderName
    """, (date_from, date_to))
    pat_pay_rows = cursor.fetchall()
    pat_pay_by_prov = {}
    for r in pat_pay_rows:
        pn  = r["ProvNum"]
        amt = float(r["PatientCollected"] or 0)
        pat_pay_by_prov[pn] = pat_pay_by_prov.get(pn, 0) + amt
    total_patient_collected = sum(pat_pay_by_prov.values())

    # Also capture negative splits (refunds) not tied to a provider
    cursor.execute("""
        SELECT COALESCE(SUM(ps.SplitAmt), 0) AS NegSplits
        FROM payment p
        INNER JOIN paysplit ps ON ps.PayNum = p.PayNum
        WHERE DATE(p.PayDate) BETWEEN %s AND %s
        AND p.PayType != 0
        AND ps.UnearnedType = 0
        AND ps.IsDiscount = 0
        AND ps.SplitAmt < 0
    """, (date_from, date_to))
    neg_row = cursor.fetchone()
    total_patient_collected += float(neg_row["NegSplits"] or 0)

    # --- Insurance collections by provider (claimproc) ---
    cursor.execute("""
        SELECT
            cp.ProvNum,
            CONCAT(prov.FName,' ',prov.LName) AS ProviderName,
            SUM(cp.InsPayAmt) AS InsCollected,
            COALESCE(ip.PlanType,'') AS PlanType,
            COALESCE(c.CarrierName,'') AS CarrierName
        FROM claimproc cp
        INNER JOIN provider prov ON prov.ProvNum = cp.ProvNum
        LEFT JOIN insplan ip ON ip.PlanNum = cp.PlanNum
        LEFT JOIN carrier c ON c.CarrierNum = ip.CarrierNum
        WHERE cp.DateCP BETWEEN %s AND %s
        AND cp.InsPayAmt != 0
        GROUP BY cp.ProvNum, ProviderName, ip.PlanType, c.CarrierName
    """, (date_from, date_to))
    ins_pay_rows = cursor.fetchall()
    ins_pay_by_prov = {}
    ins_ppo_total   = 0
    ins_med_total   = 0
    for r in ins_pay_rows:
        pn       = r["ProvNum"]
        amt      = float(r["InsCollected"] or 0)
        carrier  = r["CarrierName"] or ""
        plantype = r["PlanType"] or ""
        ins_pay_by_prov[pn] = ins_pay_by_prov.get(pn, 0) + amt
        if is_medicaid_carrier(carrier, plantype):
            ins_med_total += amt
        else:
            ins_ppo_total += amt
    total_ins_collected = sum(ins_pay_by_prov.values())
    total_collected     = total_patient_collected + total_ins_collected

    # --- Write-offs (adjustments with negative AdjAmt) ---
    cursor.execute("""
        SELECT COALESCE(SUM(ABS(AdjAmt)), 0) AS WriteOffs
        FROM adjustment
        WHERE AdjDate BETWEEN %s AND %s
        AND AdjAmt < 0
    """, (date_from, date_to))
    wo_row    = cursor.fetchone()
    # Also insurance write-offs from claimproc
    cursor.execute("""
        SELECT COALESCE(SUM(WriteOff), 0) AS InsWriteOffs
        FROM claimproc
        WHERE DateCP BETWEEN %s AND %s
        AND WriteOff > 0
    """, (date_from, date_to))
    ins_wo_row    = cursor.fetchone()
    total_writeoffs = float(wo_row["WriteOffs"] or 0) + float(ins_wo_row["InsWriteOffs"] or 0)

    # --- Medicaid collection rate ---
    # Same approach as PPO: ProcDate on both sides, (FeeBilled - WriteOff) as denominator
    cursor.execute("""
        SELECT
            COALESCE(SUM(cp.InsPayAmt), 0)                       AS MedPaid,
            COALESCE(SUM(cp.FeeBilled - cp.WriteOff), 0)         AS MedNetOwed
        FROM claimproc cp
        INNER JOIN insplan ip ON ip.PlanNum = cp.PlanNum
        LEFT JOIN carrier c ON c.CarrierNum = ip.CarrierNum
        WHERE cp.ProcDate BETWEEN %s AND %s
        AND cp.InsPayAmt > 0
        AND (ip.PlanType = 'f'
             OR LOWER(COALESCE(c.CarrierName,'')) LIKE '%keystone first%'
             OR LOWER(COALESCE(c.CarrierName,'')) LIKE '%keystone mercy%'
             OR LOWER(COALESCE(c.CarrierName,'')) LIKE '%health partner%'
             OR LOWER(COALESCE(c.CarrierName,'')) LIKE '%aetna better health%'
             OR LOWER(COALESCE(c.CarrierName,'')) LIKE '%united healthcare community%'
             OR LOWER(COALESCE(c.CarrierName,'')) LIKE '%upmc%'
             OR LOWER(COALESCE(c.CarrierName,'')) LIKE '%avesis%'
             OR LOWER(COALESCE(c.CarrierName,'')) LIKE '%dentaquest%')
    """, (date_from, date_to))
    med_rate_row  = cursor.fetchone()
    med_paid      = float(med_rate_row["MedPaid"] or 0)
    med_net_owed  = float(med_rate_row["MedNetOwed"] or 0)
    med_coll_rate = round(med_paid / med_net_owed * 100) if med_net_owed > 0 else 0

    # PPO collection rate — payments vs (billed - write-offs) from claimproc
    # Using ProcDate on both sides so numerator and denominator match the same procedures
    # This avoids the lag problem of using DateCP (payment posted date)
    cursor.execute("""
        SELECT
            COALESCE(SUM(cp.InsPayAmt), 0)                        AS PPOPaid,
            COALESCE(SUM(cp.FeeBilled - cp.WriteOff), 0)          AS PPONetOwed
        FROM claimproc cp
        INNER JOIN insplan ip ON ip.PlanNum = cp.PlanNum
        INNER JOIN carrier c  ON c.CarrierNum = ip.CarrierNum
        WHERE cp.ProcDate BETWEEN %s AND %s
        AND cp.InsPayAmt > 0
        AND ip.PlanType != 'f'
        AND LOWER(c.CarrierName) NOT LIKE '%keystone first%'
        AND LOWER(c.CarrierName) NOT LIKE '%keystone mercy%'
        AND LOWER(c.CarrierName) NOT LIKE '%health partner%'
        AND LOWER(c.CarrierName) NOT LIKE '%aetna better health%'
        AND LOWER(c.CarrierName) NOT LIKE '%united healthcare community%'
        AND LOWER(c.CarrierName) NOT LIKE '%upmc%'
        AND LOWER(c.CarrierName) NOT LIKE '%avesis%'
        AND LOWER(c.CarrierName) NOT LIKE '%dentaquest%'
    """, (date_from, date_to))
    ppo_rate_row  = cursor.fetchone()
    ppo_paid      = float(ppo_rate_row["PPOPaid"] or 0)
    ppo_net_owed  = float(ppo_rate_row["PPONetOwed"] or 0)
    ppo_coll_rate = round(ppo_paid / ppo_net_owed * 100) if ppo_net_owed > 0 else 0

    # --- Per provider summary ---
    prov_summary = []
    all_prov_nums = set(r["ProvNum"] for r in prod_rows)
    prov_names    = {r["ProvNum"]: r["ProviderName"] for r in prod_rows}
    for r in prod_rows:
        pn      = r["ProvNum"]
        prod    = float(r["Production"] or 0)
        pts     = int(r["Patients"] or 0)
        pat_c   = pat_pay_by_prov.get(pn, 0)
        ins_c   = ins_pay_by_prov.get(pn, 0)
        total_c = pat_c + ins_c
        avg_c   = round(total_c / pts) if pts > 0 else 0
        prov_summary.append({
            "name":       r["ProviderName"],
            "production": round(prod),
            "patients":   pts,
            "collected":  round(total_c),
            "avg_collection": avg_c,
        })

    # --- Outstanding AR (last 12 months) ---
    cursor.execute("""
        SELECT
            cl.ClaimNum,
            cl.DateSent,
            DATEDIFF(CURDATE(), cl.DateSent) AS DaysOut,
            c.CarrierName,
            CONCAT(pat.LName, ', ', pat.FName) AS PatientName,
            SUM(cp2.FeeBilled) - SUM(cp2.InsPayAmt) - SUM(cp2.WriteOff) AS Outstanding
        FROM claim cl
        INNER JOIN claimproc cp2 ON cp2.ClaimNum = cl.ClaimNum
        INNER JOIN patient pat ON pat.PatNum = cl.PatNum
        LEFT JOIN insplan ip2 ON ip2.PlanNum = cl.PlanNum
        LEFT JOIN carrier c ON c.CarrierNum = ip2.CarrierNum
        WHERE cl.ClaimStatus IN ('S','W')
        AND cl.ClaimType != 'PreAuth'
        AND cl.DateSent >= DATE_SUB(CURDATE(), INTERVAL 12 MONTH)
        AND cl.DateSent > '0001-01-01'
        GROUP BY cl.ClaimNum, cl.DateSent, c.CarrierName, PatientName
        HAVING Outstanding > 1
        ORDER BY Outstanding DESC
    """)
    ar_rows = cursor.fetchall()

    # Group AR by carrier
    ar_by_carrier = {}
    ar_buckets    = {"0-30": 0, "31-60": 0, "61-90": 0, "90+": 0}
    ar_counts     = {"0-30": 0, "31-60": 0, "61-90": 0, "90+": 0}

    for r in ar_rows:
        days    = int(r["DaysOut"] or 0)
        amt     = float(r["Outstanding"] or 0)
        carrier = r["CarrierName"] or "Unknown"
        pat     = r["PatientName"] or "Unknown"

        if days <= 30:   bucket = "0-30"
        elif days <= 60: bucket = "31-60"
        elif days <= 90: bucket = "61-90"
        else:            bucket = "90+"

        ar_buckets[bucket] += amt
        ar_counts[bucket]  += 1

        if carrier not in ar_by_carrier:
            ar_by_carrier[carrier] = {"total": 0, "claims": 0, "oldest": 0, "patients": []}
        ar_by_carrier[carrier]["total"]   += amt
        ar_by_carrier[carrier]["claims"]  += 1
        ar_by_carrier[carrier]["oldest"]   = max(ar_by_carrier[carrier]["oldest"], days)
        ar_by_carrier[carrier]["patients"].append({"name": pat, "days": days, "amount": round(amt)})

    # Sort carriers by total outstanding
    ar_carriers = sorted(
        [{"carrier": k, **v, "total": round(v["total"])} for k, v in ar_by_carrier.items()],
        key=lambda x: -x["total"]
    )[:15]

    # --- New patients with carrier breakdown ---
    # Step 1: Get all qualifying new patients with their PRIMARY insurance
    cursor.execute("""
        SELECT
            pat.PatNum,
            COALESCE(c.CarrierName, '') AS CarrierName,
            COALESCE(ip.PlanType, '') AS PlanType
        FROM patient pat
        LEFT JOIN patplan pp ON pp.PatNum = pat.PatNum AND pp.Ordinal = 1
        LEFT JOIN inssub ins ON ins.InsSubNum = pp.InsSubNum
        LEFT JOIN insplan ip ON ip.PlanNum = ins.PlanNum
        LEFT JOIN carrier c ON c.CarrierNum = ip.CarrierNum
        WHERE pat.DateFirstVisit BETWEEN %s AND %s
        AND pat.PatStatus = 0
        AND EXISTS (
            SELECT 1 FROM appointment a
            WHERE a.PatNum = pat.PatNum
            AND DATE(a.AptDateTime) BETWEEN %s AND %s
            AND a.AptStatus = 2
        )
    """, (date_from, date_to, date_from, date_to))
    np_rows = cursor.fetchall()

    np_ppo = np_med = np_none = np_other = 0
    np_ppo_carriers = {}
    np_med_carriers = {}
    seen_np = set()

    for r in np_rows:
        pat_num = r["PatNum"]
        if pat_num in seen_np:
            continue
        seen_np.add(pat_num)
        carrier  = r["CarrierName"] or ""
        plantype = r["PlanType"] or ""
        if not carrier and not plantype:
            np_none += 1
        elif is_medicaid_carrier(carrier, plantype):
            np_med += 1
            np_med_carriers[carrier] = np_med_carriers.get(carrier, 0) + 1
        elif plantype == "p" or carrier:
            np_ppo += 1
            np_ppo_carriers[carrier] = np_ppo_carriers.get(carrier, 0) + 1
        else:
            np_other += 1

    # Sort and take top carriers
    np_ppo_top = sorted(np_ppo_carriers.items(), key=lambda x: -x[1])[:6]
    np_med_top = sorted(np_med_carriers.items(), key=lambda x: -x[1])[:6]
    np_total   = len(seen_np)

    # --- Patient payments by carrier ---
    cursor.execute("""
        SELECT
            COALESCE(c.CarrierName, 'No Insurance') AS CarrierName,
            COUNT(DISTINCT p.PatNum) AS Patients,
            SUM(ps.SplitAmt) AS TotalPaid
        FROM payment p
        INNER JOIN paysplit ps ON ps.PayNum = p.PayNum
        INNER JOIN patient pat ON pat.PatNum = p.PatNum
        LEFT JOIN patplan pp ON pp.PatNum = p.PatNum AND pp.Ordinal = 1
        LEFT JOIN inssub ins ON ins.InsSubNum = pp.InsSubNum
        LEFT JOIN insplan ip ON ip.PlanNum = ins.PlanNum
        LEFT JOIN carrier c ON c.CarrierNum = ip.CarrierNum
        WHERE DATE(p.PayDate) BETWEEN %s AND %s
        AND p.PayAmt > 0
        AND ps.IsDiscount = 0
        AND p.PayType != 0
        AND ps.UnearnedType = 0
        GROUP BY c.CarrierName
        HAVING Patients >= 3
        ORDER BY TotalPaid DESC
        LIMIT 10
    """, (date_from, date_to))
    pay_by_carrier = cursor.fetchall()

    return {
        "date_from":          str(date_from),
        "date_to":            str(date_to),
        "total_production":   round(total_production),
        "total_collected":    round(total_collected),
        "total_procedures":   total_procedures,
        "total_patients":     total_patients,
        "patient_collected":  round(total_patient_collected),
        "ins_collected":      round(total_ins_collected),
        "total_writeoffs":    round(total_writeoffs),
        "ppo_coll_rate":      ppo_coll_rate,
        "med_coll_rate":      med_coll_rate,
        "prov_summary":       prov_summary,
        "ar_buckets":         {k: round(v) for k, v in ar_buckets.items()},
        "ar_counts":          ar_counts,
        "ar_carriers":        ar_carriers,
        "new_patients":       np_total,
        "np_ppo":             np_ppo,
        "np_med":             np_med,
        "np_none":            np_none,
        "np_other":           np_other,
        "np_ppo_carriers":    [{"name": k, "count": v} for k, v in np_ppo_top],
        "np_med_carriers":    [{"name": k, "count": v} for k, v in np_med_top],
        "pay_by_carrier":     [
            {
                "carrier":   r["CarrierName"] or "No Insurance",
                "patients":  int(r["Patients"] or 0),
                "total":     round(float(r["TotalPaid"] or 0)),
                "avg":       round(float(r["TotalPaid"] or 0) / int(r["Patients"] or 1)),
            }
            for r in pay_by_carrier
        ],
    }

@app.route("/api/financial")
@login_required
def api_financial():
    try:
        from datetime import datetime as _dt
        import mysql.connector as _mc
        date_from = request.args.get("from", _dt.now().strftime("%Y-%m-01"))
        date_to   = request.args.get("to",   _dt.now().strftime("%Y-%m-%d"))
        conn   = _mc.connect(**DB_CONFIG)
        cursor = conn.cursor(dictionary=True)
        data   = get_financial_stats(cursor, date_from, date_to)
        cursor.close()
        conn.close()
        import json as _j
        return app.response_class(
            response=_j.dumps(data, default=str),
            mimetype="application/json"
        )
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


# =============================================================================
# INSURANCE LOOKUP — Search by employer or group name
# =============================================================================

@app.route("/api/ins_lookup")
@login_required
def api_ins_lookup():
    q = (request.args.get("q") or "").strip()
    if len(q) < 2:
        return jsonify({"results": [], "query": q})
    try:
        import mysql.connector as _mc
        conn   = _mc.connect(**DB_CONFIG)
        cursor = conn.cursor(dictionary=True)
        search = f"%{q}%"

        cursor.execute("""
            SELECT
                e.EmpName                               AS MatchName,
                'Employer'                              AS MatchType,
                c.CarrierName                           AS CarrierName,
                ip.GroupName                            AS GroupName,
                ip.GroupNum                             AS GroupNum,
                COUNT(DISTINCT pat.PatNum)              AS PatientCount,
                GROUP_CONCAT(
                    DISTINCT CONCAT(pat.FName, ' ', pat.LName)
                    ORDER BY pat.LName SEPARATOR ', '
                )                                       AS SamplePatients
            FROM employer e
            INNER JOIN patient pat    ON pat.EmployerNum = e.EmployerNum
            INNER JOIN patplan pp     ON pp.PatNum = pat.PatNum AND pp.Ordinal = 1 AND pp.IsPending = 0
            INNER JOIN inssub ins     ON ins.InsSubNum = pp.InsSubNum
            INNER JOIN insplan ip     ON ip.PlanNum = ins.PlanNum
            INNER JOIN carrier c      ON c.CarrierNum = ip.CarrierNum
            WHERE e.EmpName LIKE %s AND pat.PatStatus = 0
            GROUP BY e.EmpName, c.CarrierName, ip.GroupName, ip.GroupNum
            ORDER BY PatientCount DESC LIMIT 20
        """, (search,))
        emp_rows = cursor.fetchall()

        cursor.execute("""
            SELECT
                ip.GroupName                            AS MatchName,
                'Group Name'                            AS MatchType,
                c.CarrierName                           AS CarrierName,
                ip.GroupName                            AS GroupName,
                ip.GroupNum                             AS GroupNum,
                COUNT(DISTINCT pat.PatNum)              AS PatientCount,
                GROUP_CONCAT(
                    DISTINCT CONCAT(pat.FName, ' ', pat.LName)
                    ORDER BY pat.LName SEPARATOR ', '
                )                                       AS SamplePatients
            FROM insplan ip
            INNER JOIN carrier c      ON c.CarrierNum = ip.CarrierNum
            INNER JOIN inssub ins     ON ins.PlanNum = ip.PlanNum
            INNER JOIN patplan pp     ON pp.InsSubNum = ins.InsSubNum AND pp.Ordinal = 1 AND pp.IsPending = 0
            INNER JOIN patient pat    ON pat.PatNum = pp.PatNum
            WHERE ip.GroupName LIKE %s AND ip.GroupName != '' AND pat.PatStatus = 0
            GROUP BY ip.GroupName, ip.GroupNum, c.CarrierName
            ORDER BY PatientCount DESC LIMIT 20
        """, (search,))
        grp_rows = cursor.fetchall()

        cursor.close()
        conn.close()

        seen = set()
        results = []
        for r in list(emp_rows) + list(grp_rows):
            key = (r["CarrierName"], r["GroupName"] or "", str(r["GroupNum"] or ""))
            if key in seen:
                continue
            seen.add(key)
            samples = r["SamplePatients"] or ""
            sample_list = [s.strip() for s in samples.split(",") if s.strip()]
            results.append({
                "match_name":     r["MatchName"] or "",
                "match_type":     r["MatchType"],
                "carrier":        r["CarrierName"] or "Unknown",
                "group_name":     r["GroupName"] or "",
                "group_num":      r["GroupNum"] or "",
                "patient_count":  int(r["PatientCount"] or 0),
                "sample_patients": sample_list[:5],
            })
        results.sort(key=lambda x: -x["patient_count"])
        return jsonify({"results": results[:15], "query": q})

    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

# =============================================================================
# MAIN
# =============================================================================

def main():
    log(f"=== {PRACTICE_NAME} Dashboard Starting ===")
    add_jobs()
    scheduler.start()
    log("Scheduler started.")
    update_next_runs()

    import webbrowser, time
    def open_browser():
        time.sleep(2)
        webbrowser.open("http://127.0.0.1:5000")
    threading.Thread(target=open_browser, daemon=True).start()

    print("\n" + "="*50)
    print(f"  {PRACTICE_NAME}")
    print("  Dashboard: http://127.0.0.1:5000")
    print("="*50 + "\n")

    app.run(host="0.0.0.0", port=5000, debug=False, use_reloader=False)

if __name__ == "__main__":
    main()

